# multi_source_automation.py
# Complete Updated Version with Auto Date Fix + Ultra-Permissive Entry Validation

import os
import re
import csv
import shutil
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Callable, Any, Tuple
from dataclasses import dataclass, field
from collections import defaultdict
import threading
import tkinter as tk
from tkinter import filedialog, scrolledtext, ttk, messagebox
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Attempt to import external technicians mapping (uploaded file)
try:
    import technicians  # uploaded file: technicians.py
    _TECH_REGIONS = getattr(technicians, "REGION_MAPPINGS", {})
except Exception:
    logging.warning("technicians.py not found or failed to import; continuing with empty mappings.")
    _TECH_REGIONS = {}

# ------------------------- TECHNICIAN MAPPING MODULE -------------------------
class TechnicianMapper:
    """Unified access to technician mappings across regions."""
    def __init__(self, region_mappings: Dict[str, Dict[str, str]] = None):
        self.region_mappings = region_mappings or {}
        merged = {}
        for region_name, mapping in (self.region_mappings.items() if self.region_mappings else []):
            logging.info(f"TechnicianMapper: Loading {len(mapping)} mappings from '{region_name}' region")
            for site, tech in mapping.items():
                if site not in merged:
                    merged[site] = tech
        self._map = merged
        logging.info(f"TechnicianMapper: Total {len(self._map)} unique site mappings loaded")

    def get_technician(self, site_id: str) -> str:
        if not site_id:
            logging.debug("TechnicianMapper: Empty site_id provided")
            return ""
        site_key = site_id.strip().upper()
        logging.debug(f"TechnicianMapper: Looking up '{site_key}'")
        tech = self._map.get(site_key, "")
        if tech:
            logging.debug(f"TechnicianMapper: Found '{tech}' for full ID '{site_key}'")
            return tech
        parts = site_key.split('_')
        if len(parts) > 0:
            site_num = parts[-1]
            logging.debug(f"TechnicianMapper: Trying short form '{site_num}'")
            tech = self._map.get(site_num, "")
            if tech:
                logging.debug(f"TechnicianMapper: Found '{tech}' for short form '{site_num}'")
                return tech
        logging.debug(f"TechnicianMapper: No technician found for '{site_key}'")
        return ""

TECHNICIAN_MAPPER = TechnicianMapper(_TECH_REGIONS)

# ------------------------- Configuration -------------------------
@dataclass
class ReportConfig:
    report_name: str
    excel_file: str
    sheet_name: str
    site_prefix: str
    variation_mapping: Dict[str, List[str]] = field(default_factory=dict)
    numeric_fields: List[str] = field(default_factory=list)
    allowed_sources: List[str] = field(default_factory=lambda: ["SAHARA", "MERU", "PUMA", "CCS", "TOTAL"])

    @property
    def backup_dir(self) -> Path:
        return Path(self.excel_file).parent / f"backups_{self.report_name.lower().replace(' ', '_')}"

    @property
    def unmatched_dir(self) -> Path:
        return Path(self.excel_file).parent / f"unmatched_{self.report_name.lower().replace(' ', '_')}"

class MultiSourceConfig:
    @staticmethod
    def get_new_ccs_config() -> ReportConfig:
        return ReportConfig(
            report_name="New CCS",
            excel_file=r"C:\Users\HARRISON MWEWA\Desktop\PROJECTS\practice\new ccs report.xlsx",
            sheet_name="fuel capture",
            site_prefix="IHS_CBT",
            variation_mapping={
                "CURRENT DG RUN HOURS": ["Rt", "DG Current Run Time", "GD Run Time", "Runtime", "Run Time"],
                "PREVIOUS DG RUN HOURS": ["Pre Rt", "Previous Run Time", "Previous runtime"],
                "FUEL FOUND": ["Fuel found", "Initial fuel level", "Initial"],
                "FUEL ADDED": ["Fuel added", "Added"],
                "SITE ID": ["Site ID", "cbt", "CBT"],
                "SITE NAME": ["Site name", "Site Name", "Name"],
                "SUPPLIER": ["Fuel source", "Source"],
                "DATE": ["Date"],
                "CPH": ["CPH"],
                "NAME OF TECHNICIAN": ["Technician", "Tech"]
            },
            numeric_fields=["FUEL ADDED", "FUEL FOUND", "CPH"]
        )

    @staticmethod
    def get_old_ccs_config() -> ReportConfig:
        return ReportConfig(
            report_name="Old CCS",
            excel_file=r"C:\Users\HARRISON MWEWA\Desktop\practice\old ccs report.xlsx",
            sheet_name="fuel capture",
            site_prefix="IHS_CBT",
            variation_mapping={
                "CURRENT DG RUN HOURS": ["RT", "Run Time", "GD Run Time"],
                "PREVIOUS DG RUN HOURS": ["Previous runtime", "Previous Run Time"],
                "FUEL FOUND": ["Initial", "Found", "Initial fuel level"],
                "FUEL ADDED": ["Added", "Fuel added"],
                "SITE ID": ["CBT", "Site ID"],
                "SITE NAME": ["Site name"],
                "SUPPLIER": ["Fuel source", "Source"],
                "DATE": ["Date"],
                "CPH": ["CPH"],
                "NAME OF TECHNICIAN": ["Technician"]
            },
            numeric_fields=["FUEL ADDED", "FUEL FOUND", "CPH"]
        )

    @staticmethod
    def get_nrw_config() -> ReportConfig:
        return ReportConfig(
            report_name="NRW",
            excel_file=r"C:\Users\HARRISON MWEWA\Desktop\practice\nrw report.xlsx",
            sheet_name="fuel capture",
            site_prefix="IHS_NRW",
            variation_mapping={
                "CURRENT DG RUN HOURS": ["RT", "Run Time", "GD Run Time", "DG RT"],
                "PREVIOUS DG RUN HOURS": ["Previous RT", "Previous Run Time"],
                "FUEL FOUND": ["Found", "Fuel found", "Initial fuel level"],
                "FUEL ADDED": ["Added", "Fuel added"],
                "SITE ID": ["NRW", "Site ID", "Site id"],
                "SITE NAME": ["Site name", "Site Name"],
                "SUPPLIER": ["Fuel source", "Source"],
                "DATE": ["Date"],
                "CPH": ["CPH"],
                "NAME OF TECHNICIAN": ["Technician"]
            },
            numeric_fields=["FUEL ADDED", "FUEL FOUND", "CPH"]
        )

    @staticmethod
    def get_eastern_config() -> ReportConfig:
        return ReportConfig(
            report_name="Eastern",
            excel_file=r"C:\Users\HARRISON MWEWA\Desktop\PROJECTS\practice\eastern report.xlsx",
            sheet_name="fuel capture",
            site_prefix="IHS_EST",
            variation_mapping={
                "CURRENT DG RUN HOURS": ["Run Time", "Runtime", "Rt", "Run"],
                "PREVIOUS DG RUN HOURS": ["Previous runtime"],
                "FUEL FOUND": ["Intial dp", "Initial dp", "Intial level", "Initial"],
                "FUEL ADDED": ["Added fuel", "Fuel added", "Added"],
                "SITE ID": ["EST", "Site ID", "Site I'd", "Site I'd:", "Site ID :"],
                "SITE NAME": ["Site name", "Site Name", "Name"],
                "SUPPLIER": ["source", "Fuel source"],
                "DATE": ["Date"],
                "CPH": ["CPH"],
                "NAME OF TECHNICIAN": ["Technician", "Technician Name", "TECHNICIAN NAME", "NAME OF TECHNICIAN", "Tech"]
            },
            numeric_fields=["FUEL ADDED", "FUEL FOUND", "CPH"]
        )

    @staticmethod
    def get_config(report_type: str) -> ReportConfig:
        configs = {
            'new_ccs': MultiSourceConfig.get_new_ccs_config,
            'old_ccs': MultiSourceConfig.get_old_ccs_config,
            'nrw': MultiSourceConfig.get_nrw_config,
            'eastern': MultiSourceConfig.get_eastern_config
        }
        t = report_type.lower()
        if t not in configs:
            raise ValueError(f"Unknown report type: {report_type}")
        return configs[t]()

# ------------------------- Normalizer & Parser -------------------------
class UniversalNormalizer:
    def __init__(self, config: ReportConfig):
        self.config = config
        self._setup_patterns()

    def _setup_patterns(self):
        prefix = self.config.site_prefix.split('_')[-1]
        self.site_pattern = re.compile(
            rf"(?:{prefix}|{prefix.lower()}|{prefix.title()})[\s_:,-]*([0-9]{{1,4}}[A-Z]?)",
            re.IGNORECASE
        )
        self.date_pattern = re.compile(r"(?:Date\s*[:\-=]?\s*)?(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})", re.IGNORECASE)

    def normalize_site_id(self, text: str) -> str:
        if not text:
            return ""
        m = self.site_pattern.search(text)
        if not m:
            return ""
        num = m.group(1).upper().replace("_", "")
        return f"{self.config.site_prefix}_{num}"

    def normalize_date(self, text: str) -> Optional[datetime]:
        if not text:
            return None
        m = self.date_pattern.search(text)
        if not m:
            return None
        s = m.group(1).strip().replace('-', '/')
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                continue
        try:
            return datetime.strptime(s, "%d-%b-%Y")
        except Exception:
            pass
        return None

    def normalize_supplier(self, text: str) -> str:
        if not text:
            return ""
        for src in self.config.allowed_sources:
            if re.search(rf"\b{re.escape(src)}\b", text, re.IGNORECASE):
                return src
        return ""

    def convert_numeric(self, value: Any, allow_faulty: bool = False) -> Optional[Any]:
        if value is None:
            return None
        v = str(value).strip().replace(",", "").replace("L", "").replace("l", "")
        if v == "":
            return None
        if v.lower() in ['fault', 'faulty', 'fault controller', 'faulty controller']:
            return "FAULTY" if allow_faulty else None
        if re.search(r'[a-zA-Z]', v) and re.search(r'\d', v):
            numeric_part = re.sub(r'[^0-9.]', '', v)
            if numeric_part:
                v = numeric_part
        try:
            if "." in v:
                return float(v)
            return int(v)
        except Exception:
            return "FAULTY" if allow_faulty else None

@dataclass
class ParseResult:
    entries: List[Dict]
    skipped_count: int
    unmatched_blocks: List[Dict]

class UniversalParser:
    def __init__(self, config: ReportConfig):
        self.config = config
        self.normalizer = UniversalNormalizer(config)

    def parse_file(self, file_path: str, last_processed_date: Optional[datetime] = None,
                   error_logger: Optional[Callable] = None, progress_callback: Optional[Callable] = None) -> ParseResult:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        blocks = self._split_blocks(content)
        entries = []
        unmatched = []
        skipped = 0
        total = len(blocks) if blocks else 1

        for idx, block in enumerate(blocks):
            if progress_callback:
                progress_callback(int((idx / total) * 40) + 10)

            res = self._parse_block(block, idx)
            if res is None or ('error' in res):
                skipped += 1
                reason = res.get('error', 'Parse failed') if res else 'Empty block'
                unmatched.append({'reason': reason, 'content': block})
                if error_logger:
                    error_logger(f"❌ Block {idx+1}: {reason}")
                continue

            if last_processed_date and "CURRENT VISIT DATE" in res:
                try:
                    entry_date = datetime.strptime(res["CURRENT VISIT DATE"], "%d/%m/%Y")
                    if entry_date <= last_processed_date:
                        skipped += 1
                        if error_logger:
                            error_logger(f"⏭️ Block {idx+1}: Date {res['CURRENT VISIT DATE']} is not newer than last processed date")
                        continue
                except Exception as e:
                    logging.warning(f"Could not parse date for filtering: {e}")

            entries.append(res)
            logging.info(f"✅ Parsed Block {idx+1}: key={self._make_fallback_key(res)}")

        logging.info(f"📊 Parsed {len(entries)} valid entries, skipped {skipped}")
        return ParseResult(entries=entries, skipped_count=skipped, unmatched_blocks=unmatched)

    def _split_blocks(self, content: str) -> List[str]:
        blocks = re.split(r"(?=\bDate\s*[:\-=]?\s*\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})", content, flags=re.IGNORECASE)
        if len(blocks) <= 1:
            blocks = [b.strip() for b in content.split("\n\n") if b.strip()]
        return [b.strip() for b in blocks if b.strip() and len(b.strip()) > 10]

    def _extract_site_name(self, block: str, site_id_guess: str) -> str:
        m = re.search(r"Site\s*Name\s*[:\-]?\s*(.+)", block, re.IGNORECASE)
        if m:
            name = m.group(1).strip()
            name = re.split(r"[\r\n]", name)[0].strip()
            name = re.split(r"Site\s*I'd|Site\s*ID|Run Time", name, flags=re.IGNORECASE)[0].strip()
            return name
        m2 = re.search(r"Site\s*name\s*[:\-]?\s*(.+)", block, re.IGNORECASE)
        if m2:
            return m2.group(1).splitlines()[0].strip()
        if site_id_guess:
            idx = block.find(site_id_guess)
            if idx > 0:
                prefix = block[:idx].strip().splitlines()[-1].strip()
                return prefix
        return ""

    def _make_fallback_key(self, data: Dict) -> str:
        sid = data.get("SITE ID", "")
        sname = data.get("SITE NAME", "")
        date = data.get("CURRENT VISIT DATE", "")
        if sid:
            return sid
        if sname:
            return sname.strip().upper()
        if date:
            try:
                dt = datetime.strptime(date, "%d/%m/%Y")
                return dt.strftime("%Y-%m-%d")
            except Exception:
                return date
        return ""

    def _parse_block(self, block: str, block_idx: int) -> Optional[Dict]:
        data: Dict[str, Any] = {}
        lines = [ln.strip() for ln in block.splitlines() if ln.strip()]
        if not lines:
            return None
        block_text = "\n".join(lines)

        dt = self.normalizer.normalize_date(block_text)
        if dt:
            data["CURRENT VISIT DATE"] = dt.strftime("%d/%m/%Y")

        site_id = self.normalizer.normalize_site_id(block_text)
        if site_id:
            data["SITE ID"] = site_id
            data["I.H.S SITE ID"] = site_id

        if "SITE ID" not in data:
            site_name = self._extract_site_name(block_text, site_id or "")
            if site_name:
                data["SITE NAME"] = site_name

        for line in lines:
            lower_line = line.lower()
            for col, variants in self.config.variation_mapping.items():
                if col in data and col != "NAME OF TECHNICIAN":
                    continue
                for var in variants:
                    if var.lower() in lower_line:
                        parts = re.split(r"[:\-=]", line, 1)
                        value = parts[1].strip() if len(parts) > 1 else parts[0].strip()
                        data[col] = value
                        break
                if col in data:
                    break

        if "SITE NAME" not in data:
            data["SITE NAME"] = self._extract_site_name(block_text, data.get("SITE ID", ""))

        for field in ["CURRENT DG RUN HOURS", "PREVIOUS DG RUN HOURS"]:
            if field in data:
                data[field] = self.normalizer.convert_numeric(data[field], allow_faulty=True)
        for field in self.config.numeric_fields:
            if field in data:
                data[field] = self.normalizer.convert_numeric(data[field])

        if "SUPPLIER" in data:
            data["SUPPLIER"] = self.normalizer.normalize_supplier(data["SUPPLIER"])
        else:
            data["SUPPLIER"] = self.normalizer.normalize_supplier(block_text)

        tech = ""
        site_id_for_lookup = data.get("SITE ID", "")

        logging.info(f"🔍 Looking up technician for site: '{site_id_for_lookup}'")

        if "SITE ID" in data and data["SITE ID"]:
            tech = TECHNICIAN_MAPPER.get_technician(data["SITE ID"])
            if tech:
                logging.info(f"✅ Technician '{tech}' assigned from mapping for site {data['SITE ID']}")
            else:
                logging.warning(f"⚠️ No technician found in mapping for site {data['SITE ID']}")
        else:
            logging.warning(f"⚠️ No SITE ID available for technician lookup")

        if not tech and "NAME OF TECHNICIAN" in data and data.get("NAME OF TECHNICIAN"):
            tech = data.get("NAME OF TECHNICIAN")
            logging.info(f"📋 Technician '{tech}' found in chat data")

        data["NAME OF TECHNICIAN"] = tech if tech else "N/A"
        if not tech:
            logging.error(f"❌ FINAL: No technician assigned for site {data.get('SITE ID', 'UNKNOWN')}")
        else:
            logging.info(f"✅ FINAL: Technician '{tech}' assigned for site {data.get('SITE ID', 'UNKNOWN')}")

        # ──────────────────────────────────────────────────────────────
        # ULTRA-PERMISSIVE VALIDATION - accept if ANY piece exists
        # ──────────────────────────────────────────────────────────────
        has_date      = bool(data.get("CURRENT VISIT DATE"))
        has_site_id   = bool(data.get("SITE ID"))
        has_site_name = bool(data.get("SITE NAME"))
        has_runtime   = "CURRENT DG RUN HOURS" in data and data["CURRENT DG RUN HOURS"] not in (None, "", "FAULTY")
        has_found     = "FUEL FOUND"           in data and data["FUEL FOUND"]           not in (None, "", "FAULTY")
        has_added     = "FUEL ADDED"           in data and data["FUEL ADDED"]           not in (None, "", "FAULTY")

        has_anything = (
            has_date or has_site_id or has_site_name or
            has_runtime or has_found or has_added
        )

        if not has_anything:
            return {'error': 'Completely empty block - no site, no date, no runtime, no fuel values at all'}

        return data

# ------------------------- Excel Management -------------------------
@dataclass
class SheetInfo:
    worksheet: Worksheet
    header_row: int
    headers: List[str]
    col_map: Dict[str, int]
    last_date: datetime
    existing_keys: set

class UniversalExcelManager:
    def __init__(self, config: ReportConfig):
        self.config = config
        self.backup_dir = Path(self.config.excel_file).parent / f"backups_{self.config.report_name.lower().replace(' ', '_')}"
        self.backup_dir.mkdir(parents=True, exist_ok=True)

    def backup_file(self):
        if not Path(self.config.excel_file).exists():
            logging.warning("Excel file not found for backup.")
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = self.backup_dir / f"backup_{ts}.xlsx"
        shutil.copy(self.config.excel_file, dest)
        logging.info(f"Backup created: {dest.name}")

    def load_workbook(self):
        return load_workbook(self.config.excel_file)

    def _normalize_header(self, s: Any) -> str:
        return str(s or "").strip().upper()

    def get_sheet_info(self, ws: Worksheet) -> Optional[SheetInfo]:
        header_row = None
        headers = []
        for r in ws.iter_rows(min_row=1, max_row=20):
            row_vals = [self._normalize_header(cell.value) for cell in r]
            if any("SITE ID" in v for v in row_vals) and any("CURRENT VISIT DATE" in v for v in row_vals):
                header_row = r[0].row
                headers = [str(cell.value or "").strip() for cell in r]
                break
        if header_row is None:
            return None

        col_map = {name: idx+1 for idx, name in enumerate(headers)}

        last_date = datetime.min
        date_col = None
        for name, idx in col_map.items():
            if self._normalize_header(name) == "CURRENT VISIT DATE":
                date_col = idx
                break

        if date_col:
            for rw in range(header_row+1, ws.max_row+1):
                cell = ws.cell(row=rw, column=date_col).value
                if cell is None:
                    continue
                try:
                    if isinstance(cell, datetime):
                        dt = cell
                    else:
                        dt = datetime.strptime(str(cell).strip(), "%d/%m/%Y")
                    if dt > last_date:
                        last_date = dt
                except Exception:
                    continue

        existing = set()
        return SheetInfo(worksheet=ws, header_row=header_row, headers=headers, col_map=col_map,
                        last_date=last_date, existing_keys=existing)

    def find_last_data_row(self, ws: Worksheet, col_index: int) -> int:
        for r in range(ws.max_row, 0, -1):
            v = ws.cell(row=r, column=col_index).value
            if v is not None and not (isinstance(v, str) and v.strip() == ""):
                return r
        return ws.max_row

    def ensure_column_exists(self, sheet_info: SheetInfo, column_name: str) -> SheetInfo:
        if column_name in sheet_info.headers:
            return sheet_info

        alternatives = {
            "NAME OF TECHNICIAN": ["TECHNICIAN NAME", "Technician Name", "TECHNICIAN", "Technician", "Tech Name"],
            "I.H.S SITE ID": ["IHS SITE ID", "I.H.S. SITE ID", "IHS_SITE_ID", "SITE ID"],
            "FUEL LEFT ON SITE": ["FUEL LEFT", "Fuel Left on Site", "FUEL_LEFT_ON_SITE"]
        }

        if column_name in alternatives:
            for alt_name in alternatives[column_name]:
                for idx, header in enumerate(sheet_info.headers):
                    if header and header.strip().upper() == alt_name.upper():
                        sheet_info.col_map[column_name] = idx + 1
                        logging.info(f"📋 Using existing column '{header}' for '{column_name}'")
                        return sheet_info

        logging.warning(f"⚠️ Column '{column_name}' not found, creating new column")
        new_idx = len(sheet_info.headers) + 1
        sheet_info.worksheet.cell(row=sheet_info.header_row, column=new_idx, value=column_name)
        sheet_info.headers.append(column_name)
        sheet_info.col_map[column_name] = new_idx
        return sheet_info

@dataclass
class UpdateResult:
    added: int
    faulty: int
    skipped: int

class UniversalDataUpdater:
    def __init__(self, config: ReportConfig):
        self.config = config
        self.excel_manager = UniversalExcelManager(config)

    def _fallback_key_for_entry(self, entry: Dict) -> str:
        sid = entry.get("SITE ID", "")
        sname = entry.get("SITE NAME", "")
        date = entry.get("CURRENT VISIT DATE", "")
        if sid:
            return sid
        if sname:
            return str(sname).strip().upper()
        if date:
            try:
                dt = datetime.strptime(date, "%d/%m/%Y")
                return dt.strftime("%Y-%m-%d")
            except Exception:
                return date
        return ""

    def update_excel(self, entries: List[Dict], error_logger: Optional[Callable] = None,
                     progress_callback: Optional[Callable] = None) -> UpdateResult:
        if not entries:
            logging.warning("No entries provided")
            return UpdateResult(0, 0, 0)

        self.excel_manager.backup_file()
        wb = self.excel_manager.load_workbook()
        if self.config.sheet_name not in wb.sheetnames:
            logging.error("Sheet not found")
            return UpdateResult(0, 0, 0)

        ws = wb[self.config.sheet_name]
        sheet_info = self.excel_manager.get_sheet_info(ws)
        if not sheet_info:
            logging.error("Invalid sheet format")
            return UpdateResult(0, 0, 0)

        sheet_info = self.excel_manager.ensure_column_exists(sheet_info, "NAME OF TECHNICIAN")
        sheet_info = self.excel_manager.ensure_column_exists(sheet_info, "FUEL LEFT ON SITE")
        sheet_info = self.excel_manager.ensure_column_exists(sheet_info, "I.H.S SITE ID")

        added = 0
        faulty = 0
        skipped = 0
        total = len(entries)

        site_id_col = sheet_info.col_map.get("SITE ID", 1)
        append_row = self.excel_manager.find_last_data_row(sheet_info.worksheet, site_id_col) + 1

        existing_records = set()

        date_col   = sheet_info.col_map.get("CURRENT VISIT DATE")
        site_col   = sheet_info.col_map.get("SITE ID")
        name_col   = sheet_info.col_map.get("SITE NAME")
        rt_col     = sheet_info.col_map.get("CURRENT DG RUN HOURS")
        found_col  = sheet_info.col_map.get("FUEL FOUND")
        added_col  = sheet_info.col_map.get("FUEL ADDED")

        for rw in range(sheet_info.header_row + 1, ws.max_row + 1):
            date_val = ""
            if date_col:
                v = ws.cell(row=rw, column=date_col).value
                if v:
                    if isinstance(v, datetime):
                        date_val = v.strftime("%d/%m/%Y")
                    else:
                        date_val = str(v).strip()

            site_val = ""
            if site_col:
                v = ws.cell(row=rw, column=site_col).value
                site_val = str(v).strip() if v is not None else ""
            if not site_val and name_col:
                v = ws.cell(row=rw, column=name_col).value
                site_val = str(v).strip().upper() if v is not None else ""

            rt_val    = str(ws.cell(row=rw, column=rt_col).value).strip()    if rt_col    else ""
            found_val = str(ws.cell(row=rw, column=found_col).value).strip() if found_col else ""
            added_val = str(ws.cell(row=rw, column=added_col).value).strip() if added_col else ""

            if date_val and site_val and rt_val and found_val and added_val:
                existing_records.add((date_val, site_val, rt_val, found_val, added_val))

        for idx, entry in enumerate(entries):
            if progress_callback:
                progress_callback(int((idx / total) * 40) + 60)

            entry_date    = entry.get("CURRENT VISIT DATE", "")
            entry_site    = entry.get("SITE ID", "") or entry.get("SITE NAME", "").strip().upper()
            entry_runtime = str(entry.get("CURRENT DG RUN HOURS", "")).strip()
            entry_found   = str(entry.get("FUEL FOUND", "")).strip()
            entry_added   = str(entry.get("FUEL ADDED", "")).strip()

            is_potential_duplicate = bool(entry_date and entry_site and entry_runtime and entry_found and entry_added)

            if is_potential_duplicate:
                dup_key = (entry_date, entry_site, entry_runtime, entry_found, entry_added)
                if dup_key in existing_records:
                    logging.info(f"⏭️ Skipping true duplicate: {entry_date} | {entry_site} | RT={entry_runtime} | Found={entry_found} | Added={entry_added}")
                    skipped += 1
                    continue

            row_vals = []
            for h in sheet_info.headers:
                val = ""
                if h in entry and entry[h] != "":
                    val = entry[h]
                else:
                    val = entry.get(h.upper(), entry.get(h, ""))
                    if val == "":
                        nrm = h.strip().upper()
                        if nrm == "SITE ID":
                            val = entry.get("SITE ID", "")
                        elif nrm == "I.H.S SITE ID":
                            val = entry.get("I.H.S SITE ID", entry.get("SITE ID", ""))
                        elif nrm == "SITE NAME":
                            val = entry.get("SITE NAME", "")
                        elif nrm == "CURRENT VISIT DATE":
                            val = entry.get("CURRENT VISIT DATE", "")
                        elif nrm in ("PREVIOUS DG RUN HOURS", "LAST VISIT DATE"):
                            val = entry.get("PREVIOUS DG RUN HOURS", "")
                        elif nrm == "CURRENT DG RUN HOURS":
                            val = entry.get("CURRENT DG RUN HOURS", "")
                        elif nrm == "PREVIOUS DIESEL LEVEL":
                            val = entry.get("PREVIOUS DIESEL LEVEL", "")
                        elif nrm == "FUEL FOUND":
                            val = entry.get("FUEL FOUND", "")
                        elif nrm == "FUEL ADDED":
                            val = entry.get("FUEL ADDED", "")
                        elif nrm == "FUEL LEFT ON SITE":
                            val = entry.get("FUEL LEFT ON SITE", "")
                        elif nrm in ("NAME OF TECHNICIAN", "TECHNICIAN NAME", "TECHNICIAN"):
                            val = entry.get("NAME OF TECHNICIAN", "")
                            if not val or val == "N/A":
                                site_id = entry.get("SITE ID", "")
                                if site_id:
                                    tech = TECHNICIAN_MAPPER.get_technician(site_id)
                                    if tech:
                                        val = tech
                        else:
                            val = entry.get(h.upper(), entry.get(h, ""))

                if isinstance(val, datetime):
                    val = val.strftime("%d/%m/%Y")
                row_vals.append(val)

            # Write row (WITH AUTO DATE FIX)
            for col_idx, cell_val in enumerate(row_vals, start=1):
                cell = sheet_info.worksheet.cell(row=append_row, column=col_idx, value=cell_val)

                # AUTO DATE FIX
                header_name = sheet_info.headers[col_idx - 1].strip().upper()
                if "DATE" in header_name:
                    try:
                        if isinstance(cell_val, str):
                            dt = datetime.strptime(cell_val.strip(), "%d/%m/%Y")
                            cell.value = dt
                            cell.number_format = "DD/MM/YYYY"
                    except Exception:
                        pass

            if entry_date and entry_site and entry_runtime and entry_found and entry_added:
                existing_records.add((entry_date, entry_site, entry_runtime, entry_found, entry_added))

            append_row += 1
            added += 1
            logging.info(f"✅ Added: Date={entry_date}, Site={entry_site}")

        wb.save(self.config.excel_file)
        logging.info(f"Excel saved. Added={added} | Faulty={faulty} | Skipped={skipped}")
        return UpdateResult(added=added, faulty=faulty, skipped=skipped)

# ------------------------- Propagation & Unmatched Exporter -------------------------
@dataclass
class PropagationResult:
    sites_processed: int
    rows_updated: int
    dates_filled: int
    diesel_levels_filled: int

class UniversalDataPropagator:
    def __init__(self, config: ReportConfig):
        self.config = config
        self.excel_manager = UniversalExcelManager(config)

    def _fallback_key_from_row(self, ws: Worksheet, row: int, col_map: Dict[str, int]) -> str:
        sid_col = col_map.get("SITE ID")
        sname_col = col_map.get("SITE NAME")
        curr_date_col = None
        last_visit_col = None
        for name, idx in col_map.items():
            nrm = str(name).strip().upper()
            if nrm == "CURRENT VISIT DATE":
                curr_date_col = idx
            if nrm == "LAST VISIT DATE":
                last_visit_col = idx
        sid = ws.cell(row=row, column=sid_col).value if sid_col else None
        sname = ws.cell(row=row, column=sname_col).value if sname_col else None
        date_val = None
        if last_visit_col:
            date_val = ws.cell(row=row, column=last_visit_col).value
        if not date_val and curr_date_col:
            date_val = ws.cell(row=row, column=curr_date_col).value
        if sid and str(sid).strip():
            return str(sid).strip()
        if sname and str(sname).strip():
            return str(sname).strip().upper()
        if date_val:
            if isinstance(date_val, datetime):
                return date_val.strftime("%Y-%m-%d")
            try:
                dt = datetime.strptime(str(date_val).strip(), "%d/%m/%Y")
                return dt.strftime("%Y-%m-%d")
            except Exception:
                return str(date_val).strip()
        return ""

    def propagate_data(self, progress_callback: Optional[Callable] = None) -> PropagationResult:
        logging.info("🔄 Propagation started (using LAST VISIT DATE as previous date)...")
        self.excel_manager.backup_file()
        wb = self.excel_manager.load_workbook()
        if self.config.sheet_name not in wb.sheetnames:
            logging.error("Sheet not found")
            return PropagationResult(0,0,0,0)
        ws = wb[self.config.sheet_name]
        sheet_info = self.excel_manager.get_sheet_info(ws)
        if not sheet_info:
            logging.error("Invalid sheet")
            return PropagationResult(0,0,0,0)

        groups = defaultdict(list)
        for r in range(sheet_info.header_row+1, ws.max_row+1):
            fk = self._fallback_key_from_row(ws, r, sheet_info.col_map)
            if fk:
                groups[fk].append(r)

        sites_processed = 0
        rows_updated = 0
        dates_filled = 0
        diesel_filled = 0
        total_sites = len(groups) if groups else 1

        col_prev_date = None
        col_curr_date = None
        col_prev_diesel = None
        col_fuel_left = None
        for name, idx in sheet_info.col_map.items():
            nrm = str(name).strip().upper()
            if nrm == "LAST VISIT DATE":
                col_prev_date = idx
            elif nrm == "CURRENT VISIT DATE":
                col_curr_date = idx
            elif nrm == "PREVIOUS DIESEL LEVEL":
                col_prev_diesel = idx
            elif nrm == "FUEL LEFT ON SITE":
                col_fuel_left = idx

        for idx, (fk, rows) in enumerate(groups.items()):
            if progress_callback:
                progress_callback(int((idx / total_sites) * 30) + 60)
            rows.sort()
            prev_curr_date = None
            prev_fuel_left = None
            for r in rows:
                modified = False
                if col_prev_date and col_curr_date:
                    prev_val = ws.cell(row=r, column=col_prev_date).value
                    curr_val = ws.cell(row=r, column=col_curr_date).value
                    if (prev_val is None or (isinstance(prev_val, str) and not prev_val.strip())):
                        if prev_curr_date:
                            ws.cell(row=r, column=col_prev_date, value=prev_curr_date)
                            dates_filled += 1
                            modified = True
                        elif curr_val and not (isinstance(curr_val, str) and not curr_val.strip()):
                            ws.cell(row=r, column=col_prev_date, value=curr_val)
                            dates_filled += 1
                            prev_curr_date = curr_val
                            modified = True
                    else:
                        prev_curr_date = prev_val
                if col_prev_diesel and col_fuel_left:
                    prev_diesel = ws.cell(row=r, column=col_prev_diesel).value
                    fuel_left = ws.cell(row=r, column=col_fuel_left).value
                    if (prev_diesel is None or (isinstance(prev_diesel, str) and not prev_diesel.strip())):
                        if prev_fuel_left:
                            ws.cell(row=r, column=col_prev_diesel, value=prev_fuel_left)
                            diesel_filled += 1
                            modified = True
                        elif fuel_left and not (isinstance(fuel_left, str) and not fuel_left.strip()):
                            ws.cell(row=r, column=col_prev_diesel, value=fuel_left)
                            diesel_filled += 1
                            prev_fuel_left = fuel_left
                            modified = True
                    else:
                        prev_fuel_left = prev_diesel
                if modified:
                    rows_updated += 1
            sites_processed += 1

        wb.save(self.config.excel_file)
        logging.info(f"Propagation finished: sites={sites_processed}, rows_updated={rows_updated}")
        return PropagationResult(sites_processed, rows_updated, dates_filled, diesel_filled)

class UnmatchedBlockExporter:
    def __init__(self, config: ReportConfig):
        self.outdir = Path(config.excel_file).parent / f"unmatched_{config.report_name.lower().replace(' ', '_')}"
        self.outdir.mkdir(parents=True, exist_ok=True)

    def export(self, unmatched_blocks: List[Dict]):
        if not unmatched_blocks:
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = self.outdir / f"unmatched_{ts}.csv"
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["BlockNumber", "Reason", "Preview", "Content"])
            for i, b in enumerate(unmatched_blocks, 1):
                preview = b['content'][:120].replace("\n", " ")
                w.writerow([i, b.get('reason', ''), preview, b.get('content', '')])
        logging.info(f"Unmatched exported: {path.name}")

# ------------------------- Controller -------------------------
class MultiSourceController:
    def __init__(self, report_type: str, excel_file: Optional[str] = None):
        self.report_type = report_type.lower()
        self.config = MultiSourceConfig.get_config(self.report_type)
        if excel_file:
            self.config.excel_file = excel_file
        self.parser = UniversalParser(self.config)
        self.updater = UniversalDataUpdater(self.config)
        self.propagator = UniversalDataPropagator(self.config)
        self.exporter = UnmatchedBlockExporter(self.config)

    def run_automation(self, chat_file: str, error_logger: Optional[Callable] = None,
                       progress_callback: Optional[Callable] = None) -> Tuple[UpdateResult, int]:
        logging.info(f"Starting automation for {self.config.report_name} ...")
        last_date = None
        try:
            wb = self.updater.excel_manager.load_workbook()
            if self.config.sheet_name in wb.sheetnames:
                ws = wb[self.config.sheet_name]
                sheet_info = self.updater.excel_manager.get_sheet_info(ws)
                if sheet_info:
                    last_date = sheet_info.last_date
                    if last_date and last_date != datetime.min:
                        logging.info(f"📅 Last processed date in Excel: {last_date.strftime('%d/%m/%Y')}")
                        logging.info(f"✨ Will only process entries AFTER this date")
        except Exception as e:
            logging.warning(f"Could not determine last date: {e}")

        parse_result = self.parser.parse_file(
            chat_file,
            last_processed_date=last_date,
            error_logger=error_logger,
            progress_callback=progress_callback
        )

        if not parse_result.entries:
            logging.warning("No new entries to process after filtering by date")
            return UpdateResult(0, 0, 0), parse_result.skipped_count

        if parse_result.unmatched_blocks:
            self.exporter.export(parse_result.unmatched_blocks)

        result = self.updater.update_excel(parse_result.entries, error_logger, progress_callback)
        return result, parse_result.skipped_count

    def run_propagation(self, progress_callback: Optional[Callable] = None) -> PropagationResult:
        return self.propagator.propagate_data(progress_callback)

# ------------------------- GUI -------------------------
class MultiSourceGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Multi-Source Automation (Updated)")
        self.root.geometry("1100x800")
        self.selected_report = tk.StringVar(value="eastern")
        self.controller = None
        self.chat_file = None
        self.excel_file = None

        self._setup_ui()
        self._setup_logging()
        self._update_controller()

    def _setup_ui(self):
        main = ttk.Frame(self.root, padding=10)
        main.pack(fill='both', expand=True)
        rf = ttk.LabelFrame(main, text="Report Type")
        rf.pack(fill='x', padx=5, pady=5)
        for text, val in [("New CCS", "new_ccs"), ("Old CCS", "old_ccs"), ("NRW", "nrw"), ("Eastern", "eastern")]:
            ttk.Radiobutton(rf, text=text, variable=self.selected_report, value=val, command=self._update_controller).pack(side='left', padx=8)

        ef = ttk.Frame(main)
        ef.pack(fill='x', padx=5, pady=5)
        self.excel_label = ttk.Label(ef, text="Excel file: Not selected (using default)")
        self.excel_label.pack(side='left', padx=5)
        ttk.Button(ef, text="Select Excel File", command=self.select_excel_file).pack(side='right', padx=5)

        ff = ttk.Frame(main)
        ff.pack(fill='x', padx=5, pady=5)
        self.file_label = ttk.Label(ff, text="No chat file selected")
        self.file_label.pack(side='left', padx=5)
        ttk.Button(ff, text="Select Chat File", command=self.select_file).pack(side='right', padx=5)

        lf = ttk.LabelFrame(main, text="Log")
        lf.pack(fill='both', expand=True, padx=5, pady=5)
        self.text_area = scrolledtext.ScrolledText(lf, state='disabled', height=20, font=("Consolas", 10))
        self.text_area.pack(fill='both', expand=True)

        ef = ttk.LabelFrame(main, text="Errors")
        ef.pack(fill='x', padx=5, pady=5)
        self.error_area = scrolledtext.ScrolledText(ef, state='disabled', height=6, font=("Consolas", 10))
        self.error_area.pack(fill='x', expand=True)

        bf = ttk.Frame(main)
        bf.pack(fill='x', padx=5, pady=5)
        self.start_btn = ttk.Button(bf, text="Start Automation", command=self.start_automation_thread, state='disabled')
        self.start_btn.pack(side='left', padx=5)
        self.propagate_btn = ttk.Button(bf, text="Propagate Data", command=self.start_propagation_thread)
        self.propagate_btn.pack(side='left', padx=5)
        ttk.Button(bf, text="Clear Logs", command=self.clear_logs).pack(side='left', padx=5)
        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(bf, textvariable=self.status_var).pack(side='right')

        self.progress = ttk.Progressbar(main, length=800, mode='determinate')
        self.progress.pack(fill='x', padx=5, pady=5)

    def _setup_logging(self):
        logger = logging.getLogger()
        logger.setLevel(logging.INFO)
        for h in logger.handlers[:]:
            logger.removeHandler(h)

        class GUIHandler(logging.Handler):
            def __init__(self, text_widget):
                super().__init__()
                self.text_widget = text_widget
            def emit(self, record):
                try:
                    msg = self.format(record)
                except Exception:
                    msg = str(record)
                def append():
                    try:
                        self.text_widget.configure(state='normal')
                        self.text_widget.insert(tk.END, msg + "\n")
                        self.text_widget.see(tk.END)
                        self.text_widget.configure(state='disabled')
                    except tk.TclError:
                        pass
                try:
                    self.text_widget.after(0, append)
                except Exception:
                    append()

        handler = GUIHandler(self.text_area)
        handler.setFormatter(logging.Formatter('%(asctime)s - %(message)s', datefmt='%H:%M:%S'))
        logging.getLogger().addHandler(handler)

    def _update_controller(self):
        rt = self.selected_report.get()
        self.controller = MultiSourceController(rt, excel_file=self.excel_file)
        logging.info(f"Switched to {self.controller.config.report_name}")
        self.status_var.set(f"Ready - {self.controller.config.report_name}")

    def select_excel_file(self):
        p = filedialog.askopenfilename(
            title="Select Excel Report File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if p:
            self.excel_file = p
            self.excel_label.config(text=f"Excel file: {os.path.basename(p)}")
            logging.info(f"Excel file selected: {os.path.basename(p)}")
            self._update_controller()

    def select_file(self):
        p = filedialog.askopenfilename(filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if p:
            self.chat_file = p
            self.file_label.config(text=os.path.basename(p))
            self.start_btn.config(state='normal')
            logging.info(f"Chat file selected: {os.path.basename(p)}")

    def clear_logs(self):
        self.text_area.configure(state='normal')
        self.text_area.delete('1.0', tk.END)
        self.text_area.configure(state='disabled')
        self.error_area.configure(state='normal')
        self.error_area.delete('1.0', tk.END)
        self.error_area.configure(state='disabled')
        logging.info("Logs cleared")

    def log_error(self, msg: str):
        self.error_area.configure(state='normal')
        self.error_area.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {msg}\n")
        self.error_area.see(tk.END)
        self.error_area.configure(state='disabled')

    def update_progress(self, v: int):
        self.progress['value'] = v
        self.root.update_idletasks()

    def run_automation(self):
        if not self.chat_file:
            logging.error("No chat file selected")
            return
        try:
            self.status_var.set("Processing...")
            self.update_progress(5)
            result, skipped = self.controller.run_automation(self.chat_file, error_logger=self.log_error, progress_callback=self.update_progress)
            self.update_progress(100)
            msg = f"Completed: Added={result.added}, Faulty={result.faulty}, Skipped={result.skipped + skipped}"
            logging.info(msg)
            messagebox.showinfo("Done", msg)
        except Exception as e:
            logging.exception("Error during automation")
            messagebox.showerror("Error", str(e))
        finally:
            self.update_progress(0)
            self.start_btn.config(state='normal')
            self.propagate_btn.config(state='normal')
            self.status_var.set("Ready")

    def run_propagation(self):
        try:
            self.status_var.set("Propagating...")
            self.update_progress(5)
            res = self.controller.run_propagation(progress_callback=self.update_progress)
            self.update_progress(100)
            messagebox.showinfo("Propagation", f"Sites processed: {res.sites_processed}, rows updated: {res.rows_updated}")
            logging.info(f"Propagation done: {res.sites_processed} sites")
        except Exception as e:
            logging.exception("Propagation error")
            messagebox.showerror("Error", str(e))
        finally:
            self.update_progress(0)
            self.start_btn.config(state='normal')
            self.propagate_btn.config(state='normal')
            self.status_var.set("Ready")

    def start_automation_thread(self):
        self.clear_logs()
        self.start_btn.config(state='disabled')
        self.propagate_btn.config(state='disabled')
        threading.Thread(target=self.run_automation, daemon=True).start()

    def start_propagation_thread(self):
        self.clear_logs()
        self.start_btn.config(state='disabled')
        self.propagate_btn.config(state='disabled')
        threading.Thread(target=self.run_propagation, daemon=True).start()

def main():
    root = tk.Tk()
    try:
        ttk.Style().theme_use('clam')
    except Exception:
        pass
    app = MultiSourceGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()