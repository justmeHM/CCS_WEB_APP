"""
WhatsApp to Excel Automation Tool - Modular Architecture
Author: Enhanced Version
Description: Parses WhatsApp fuel refueling messages and updates Excel sheets
Version: 2.1 - Added Data Propagation Module
"""

import os
import re
import csv
import shutil
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Callable
from dataclasses import dataclass, field
from collections import defaultdict
import threading
import tkinter as tk
from tkinter import filedialog, scrolledtext, ttk, messagebox
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ============================================================================
# CONFIGURATION MODULE
# ============================================================================

@dataclass
class Config:
    """Application configuration"""
    excel_file: str = ""
    fuel_sheet: str = "fuel capture"
    ms_sheet: str = "MS CAPTURE"
    
    variation_mapping: Dict[str, List[str]] = field(default_factory=lambda: {
        "CURRENT DG RUN HOURS": ["Rt", "DG Current Run Time", "GD Run Time", "GD Run", "Runtime", "runtime"],
        "PREVIOUS DG RUN HOURS": ["Pre Rt", "Previous Run Time", "Previous", "Previous Run time", "Previous runtime"],
        "FUEL FOUND": ["Fuel found", "Initial fuel level", "Initial", "Fuel Total", "initial"],
        "FUEL ADDED": ["Fuel added", "Fuel Added", "Added", "fuel added"],
        "SITE ID": ["Site ID", "cbt", "CBT", "Site id", "site id", "SiteID", "Site Id"],
        "SITE NAME": ["Site name", "Site Name", "Name", "site name"],
        "SUPPLIER": ["Fuel source", "Source"],
        "DATE": ["Date"],
        "CPH": ["CPH"],
        "NAME OF TECHNICIAN": ["Technician", "Tech"]
    })
    
    numeric_fields: List[str] = field(default_factory=lambda: ["FUEL ADDED", "FUEL FOUND", "CPH"])
    allowed_sources: List[str] = field(default_factory=lambda: ["SAHARA", "MERU", "PUMA", "CCS FUEL", "TOTAL"])
    
    @property
    def backup_dir(self) -> Path:
        if not self.excel_file:
            return Path.cwd() / "backups"
        return Path(self.excel_file).parent / "backups"
    
    @property
    def unmatched_dir(self) -> Path:
        if not self.excel_file:
            return Path.cwd() / "unmatched_blocks"
        return Path(self.excel_file).parent / "unmatched_blocks"


# ============================================================================
# TECHNICIAN MAPPING MODULE
# ============================================================================

class TechnicianMapper:
    """Handles site to technician mapping"""
    
    SITE_TECHNICIAN_MAP = {
    'KAMBOLE': [
        'IHS_CBT_002M', 'IHS_CBT_005M', 'IHS_CBT_011M', 'IHS_CBT_013M',
        'IHS_CBT_015M', 'IHS_CBT_017M', 'IHS_CBT_023M', 'IHS_CBT_025M',
        'IHS_CBT_029M', 'IHS_CBT_031M', 'IHS_CBT_037M', 'IHS_CBT_041M',
        'IHS_CBT_072M', 'IHS_CBT_216M', 'IHS_CBT_218A', 'IHS_CBT_219M',
        'IHS_CBT_220M', 'IHS_CBT_222M', 'IHS_CBT_226A', 'IHS_CBT_226M',
        'IHS_CBT_233A', 'IHS_CBT_245A', 'IHS_CBT_257A', 'IHS_CBT_264A',
        'IHS_CBT_291A', 'IHS_CBT_335A', 'IHS_CBT_344A', 'IHS_CBT_347A',
        'IHS_CBT_359A', 'IHS_CBT_360A', 'IHS_CBT_361A', 'IHS_CBT_364A',
        'IHS_CBT_378A', 'IHS_CBT_390A'
    ],

    'ROYD': [
        'IHS_CBT_001M', 'IHS_CBT_003M', 'IHS_CBT_004M', 'IHS_CBT_006M',
        'IHS_CBT_007M', 'IHS_CBT_008M', 'IHS_CBT_009M', 'IHS_CBT_010M',
        'IHS_CBT_012M', 'IHS_CBT_014M', 'IHS_CBT_016M', 'IHS_CBT_020M',
        'IHS_CBT_021M', 'IHS_CBT_022M', 'IHS_CBT_024M', 'IHS_CBT_027M',
        'IHS_CBT_032M', 'IHS_CBT_033M', 'IHS_CBT_034M', 'IHS_CBT_035M',
        'IHS_CBT_036M', 'IHS_CBT_156M', 'IHS_CBT_164M', 'IHS_CBT_207A',
        'IHS_CBT_214A', 'IHS_CBT_224M', 'IHS_CBT_232A', 'IHS_CBT_246A',
        'IHS_CBT_252A', 'IHS_CBT_253A', 'IHS_CBT_283A', 'IHS_CBT_288A',
        'IHS_CBT_293A', 'IHS_CBT_325A', 'IHS_CBT_331A', 'IHS_CBT_332A',
        'IHS_CBT_338A', 'IHS_CBT_349A', 'IHS_CBT_362A', 'IHS_CBT_363A',
        'IHS_CBT_385A'
    ],

    'FACKSON': [
        'IHS_CBT_061M', 'IHS_CBT_073M', 'IHS_CBT_074M', 'IHS_CBT_075M',
        'IHS_CBT_081M', 'IHS_CBT_082M', 'IHS_CBT_083M', 'IHS_CBT_088M',
        'IHS_CBT_092M', 'IHS_CBT_094M', 'IHS_CBT_096M', 'IHS_CBT_102M',
        'IHS_CBT_104M', 'IHS_CBT_105M', 'IHS_CBT_108M', 'IHS_CBT_217A',
        'IHS_CBT_218M', 'IHS_CBT_222A', 'IHS_CBT_223M', 'IHS_CBT_237A',
        'IHS_CBT_244A', 'IHS_CBT_251A', 'IHS_CBT_260A', 'IHS_CBT_274A',
        'IHS_CBT_289A', 'IHS_CBT_297A', 'IHS_CBT_299A', 'IHS_CBT_300A',
        'IHS_CBT_305A', 'IHS_CBT_334A', 'IHS_CBT_345A', 'IHS_CBT_353A',
        'IHS_CBT_355A', 'IHS_CBT_356A', 'IHS_CBT_373A'
    ],

    'JAULA': [
        'IHS_CBT_080M', 'IHS_CBT_087M', 'IHS_CBT_089M', 'IHS_CBT_090M',
        'IHS_CBT_093M', 'IHS_CBT_098M', 'IHS_CBT_099M', 'IHS_CBT_103M',
        'IHS_CBT_111M', 'IHS_CBT_112M', 'IHS_CBT_113M', 'IHS_CBT_114M',
        'IHS_CBT_115M', 'IHS_CBT_116M', 'IHS_CBT_120M', 'IHS_CBT_217M',
        'IHS_CBT_220A', 'IHS_CBT_228A', 'IHS_CBT_229A', 'IHS_CBT_235M',
        'IHS_CBT_241A', 'IHS_CBT_249A', 'IHS_CBT_263A', 'IHS_CBT_266A',
        'IHS_CBT_269A', 'IHS_CBT_276A', 'IHS_CBT_285A', 'IHS_CBT_296A',
        'IHS_CBT_333A', 'IHS_CBT_351A', 'IHS_CBT_354A', 'IHS_CBT_381A',
        'IHS_CBT_383A', 'IHS_CBT_384A', 'IHS_CBT_386A', 'IHS_CBT_393A'
    ],

    'ISAAC': [
        'IHS_CBT_042M', 'IHS_CBT_043M', 'IHS_CBT_044M', 'IHS_CBT_047M',
        'IHS_CBT_048M', 'IHS_CBT_049M', 'IHS_CBT_050M', 'IHS_CBT_052M',
        'IHS_CBT_054M', 'IHS_CBT_057M', 'IHS_CBT_058M', 'IHS_CBT_062M',
        'IHS_CBT_064M', 'IHS_CBT_065M', 'IHS_CBT_068M', 'IHS_CBT_069M',
        'IHS_CBT_076M', 'IHS_CBT_137M', 'IHS_CBT_161M', 'IHS_CBT_172M',
        'IHS_CBT_179M', 'IHS_CBT_193M', 'IHS_CBT_201A', 'IHS_CBT_210A',
        'IHS_CBT_210M', 'IHS_CBT_211A', 'IHS_CBT_227M', 'IHS_CBT_230A',
        'IHS_CBT_233M', 'IHS_CBT_254A', 'IHS_CBT_265A', 'IHS_CBT_277A',
        'IHS_CBT_278A', 'IHS_CBT_307A', 'IHS_CBT_312A', 'IHS_CBT_314A',
        'IHS_CBT_339A', 'IHS_CBT_343A', 'IHS_CBT_367A', 'IHS_CBT_369A',
        'IHS_CBT_377A'
    ],

    'JUSTIN': [
        'IHS_CBT_040M', 'IHS_CBT_045M', 'IHS_CBT_056M', 'IHS_CBT_060M',
        'IHS_CBT_070M', 'IHS_CBT_085M', 'IHS_CBT_091M', 'IHS_CBT_148M',
        'IHS_CBT_227A', 'IHS_CBT_234A', 'IHS_CBT_235A', 'IHS_CBT_236A',
        'IHS_CBT_306A', 'IHS_CBT_309A', 'IHS_CBT_320A', 'IHS_CBT_326A',
        'IHS_CBT_346A'
    ],

    'SAMUEL': [
        'IHS_CBT_078M', 'IHS_CBT_079M', 'IHS_CBT_084M', 'IHS_CBT_107M',
        'IHS_CBT_117M', 'IHS_CBT_118M', 'IHS_CBT_119M', 'IHS_CBT_121M',
        'IHS_CBT_126M', 'IHS_CBT_127M', 'IHS_CBT_129M', 'IHS_CBT_132M',
        'IHS_CBT_133M', 'IHS_CBT_134M', 'IHS_CBT_146M', 'IHS_CBT_167M',
        'IHS_CBT_206A', 'IHS_CBT_225M', 'IHS_CBT_238M', 'IHS_CBT_239A',
        'IHS_CBT_241M', 'IHS_CBT_242A', 'IHS_CBT_243A', 'IHS_CBT_250A',
        'IHS_CBT_259A', 'IHS_CBT_261A', 'IHS_CBT_267A', 'IHS_CBT_301A',
        'IHS_CBT_310A', 'IHS_CBT_342A', 'IHS_CBT_352A', 'IHS_CBT_368A',
        'IHS_CBT_370A', 'IHS_CBT_371A', 'IHS_CBT_375A', 'IHS_CBT_380A'
    ],

    'SUNDAY': [
        'IHS_CBT_181M', 'IHS_CBT_215M', 'IHS_CBT_271M', 'IHS_CBT_279A',
        'IHS_CBT_319A'
    ],

    'DAVID': [
        'IHS_CBT_155M', 'IHS_CBT_203M', 'IHS_CBT_213M', 'IHS_CBT_223A',
        'IHS_CBT_328A', 'IHS_CBT_329A'
    ],

    'KENNEDY': ['IHS_CBT_018M'],

    'DALITSO': ['IHS_CBT_162M'],

    'PATRON': []  # had some earlier but overwritten in final mapping
}

    @classmethod
    def get_technician(cls, site_id: str) -> str:
        """Get technician name for a site ID"""
        return cls.SITE_TECHNICIAN_MAP.get(site_id, "N/A")


# ============================================================================
# DATA NORMALIZATION MODULE
# ============================================================================

class DataNormalizer:
    """Handles data normalization and validation"""
    
    SITE_ID_PATTERN = re.compile(r"CBT[_\s-]?\d+[A-Z]?", re.IGNORECASE)
    DATE_PATTERN = re.compile(r'Date\s*[:\-]?\s*(\d{1,2}/\d{1,2}/\d{2,4})', re.IGNORECASE)
    
    def __init__(self, config: Config):
        self.config = config
    
    def normalize_site_id(self, text: str) -> str:
        """Extract and normalize site ID from text"""
        if not text:
            return ""
        
        match = self.SITE_ID_PATTERN.search(text.upper())
        if not match:
            return ""
        
        site_id = match.group(0).replace(" ", "_").replace("-", "_")
        
        if not site_id.startswith("IHS_"):
            site_id = f"IHS_{site_id}"
        
        site_id = site_id.replace("IHSCBT", "IHS_CBT")
        if "CBT" in site_id and "CBT_" not in site_id:
            site_id = site_id.replace("CBT", "CBT_")
        
        return site_id
    
    def normalize_date(self, text: str) -> Optional[datetime]:
        """Extract and normalize date from text"""
        if not text:
            return None
        
        match = self.DATE_PATTERN.search(text)
        if not match:
            return None
        
        date_str = match.group(1).strip()
        for fmt in ("%d/%m/%y", "%d/%m/%Y"):
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        
        return None
    
    def normalize_supplier(self, text: str) -> str:
        """Extract and normalize supplier from text"""
        if not text:
            return ""
        
        for source in self.config.allowed_sources:
            if re.search(rf"\b{source}\b", text, re.IGNORECASE):
                return source
        
        return ""
    
    def convert_numeric(self, value: str, allow_faulty: bool = False) -> Optional[any]:
        """Convert string to numeric value"""
        if not value:
            return None
        
        value = str(value).replace(",", "").strip()
        
        try:
            return float(value) if "." in value else int(value)
        except ValueError:
            return "FAULTY" if allow_faulty else None


# ============================================================================
# MESSAGE PARSER MODULE
# ============================================================================

@dataclass
class ParseResult:
    """Result of message parsing"""
    entries: List[Dict]
    skipped_count: int
    unmatched_blocks: List[Dict]


class MessageParser:
    """Parses WhatsApp messages into structured data"""
    
    REFUEL_SPLIT = re.compile(r"REFUELING TEMPLATE", re.IGNORECASE)
    
    def __init__(self, config: Config):
        self.config = config
        self.normalizer = DataNormalizer(config)
        self.technician_mapper = TechnicianMapper()
    
    def parse_file(self, 
                   file_path: str, 
                   error_logger: Optional[Callable] = None,
                   progress_callback: Optional[Callable] = None) -> ParseResult:
        """Parse WhatsApp chat file"""
        
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        
        raw_blocks = self.REFUEL_SPLIT.split(content)[1:]
        entries = []
        unmatched_blocks = []
        skipped_count = 0
        total_blocks = len(raw_blocks)
        
        for block_idx, block in enumerate(raw_blocks):
            if progress_callback:
                progress = int((block_idx / total_blocks) * 40) + 20
                progress_callback(progress)
            
            result = self._parse_block(block, block_idx)
            
            if result is None:
                skipped_count += 1
                unmatched_blocks.append({
                    'reason': 'Parse failed',
                    'content': block
                })
                if error_logger:
                    error_logger(f"❌ Block {block_idx + 1}: Parse failed")
                continue
            
            if 'error' in result:
                skipped_count += 1
                unmatched_blocks.append({
                    'reason': result['error'],
                    'content': block
                })
                if error_logger:
                    error_logger(f"❌ Block {block_idx + 1}: {result['error']}")
                continue
            
            entries.append(result)
            logging.info(f"✅ Block {block_idx + 1}: Parsed {result.get('SITE ID', 'Unknown')}")
        
        logging.info(f"📊 Summary: {len(entries)} valid entries parsed, {skipped_count} skipped")
        
        return ParseResult(
            entries=entries,
            skipped_count=skipped_count,
            unmatched_blocks=unmatched_blocks
        )
    
    def _parse_block(self, block: str, block_idx: int) -> Optional[Dict]:
        """Parse a single message block"""
        data = {}
        lines = [line.strip() for line in block.split("\n") if line.strip()]
        
        if not lines:
            return None
        
        block_text = "\n".join(lines)
        
        # Extract fields using variation mapping
        for line in lines:
            lower_line = line.lower()
            for col, variants in self.config.variation_mapping.items():
                if col in data and col != "NAME OF TECHNICIAN":
                    continue
                
                for var in variants:
                    if var.lower() in lower_line:
                        parts = re.split(r"[:\-]", line, 1)
                        value = parts[1].strip() if len(parts) > 1 else parts[0].strip()
                        data[col] = value
                        break
                
                if col in data:
                    break
        
        # Normalize site ID
        raw_site = data.get("SITE ID", "") or block_text
        normalized_site = self.normalizer.normalize_site_id(raw_site)
        
        # More flexible validation: entry is valid if it has Site ID OR Date OR Runtime
        # Normalize date
        dt = self.normalizer.normalize_date(block_text)
        
        # Check if we have at least one of: Site ID, Date, or Runtime
        has_site_id = bool(normalized_site)
        has_date = bool(dt)
        has_runtime = bool(data.get("CURRENT DG RUN HOURS"))
        
        if not (has_site_id or has_date or has_runtime):
            return {'error': 'Missing all required fields (need at least Site ID, Date, or Runtime)'}
        
        # Set Site ID (even if empty, we'll allow it now)
        data["SITE ID"] = normalized_site if normalized_site else ""
        
        # Map technician (only if we have a site ID)
        if normalized_site:
            data["NAME OF TECHNICIAN"] = self.technician_mapper.get_technician(normalized_site)
        
        # Set date (even if empty, we'll allow it now)
        if dt:
            data["CURRENT VISIT DATE"] = dt.strftime("%d/%m/%Y")
        else:
            data["CURRENT VISIT DATE"] = ""
        
        # Normalize supplier
        data["SUPPLIER"] = self.normalizer.normalize_supplier(data.get("SUPPLIER", block_text))
        data["SITE NAME"] = data.get("SITE NAME", "").strip()
        
        # Convert numeric fields
        for field in ["CURRENT DG RUN HOURS", "PREVIOUS DG RUN HOURS"]:
            if field in data:
                data[field] = self.normalizer.convert_numeric(data[field], allow_faulty=True)
        
        for field in self.config.numeric_fields:
            if field in data:
                data[field] = self.normalizer.convert_numeric(data[field])
        
        return data


# ============================================================================
# EXCEL MANAGER MODULE
# ============================================================================

@dataclass
class SheetInfo:
    """Information about an Excel sheet"""
    worksheet: Worksheet
    header_row: int
    headers: List[str]
    col_map: Dict[str, int]
    last_date: datetime
    existing_keys: set


class ExcelManager:
    """Manages Excel file operations"""
    
    def __init__(self, config: Config):
        self.config = config
        self.backup_dir = config.backup_dir
        self.backup_dir.mkdir(exist_ok=True)
    
    def backup_file(self):
        """Create a backup of the Excel file"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.backup_dir / f"backup_{timestamp}.xlsx"
        shutil.copy(self.config.excel_file, backup_path)
        logging.info(f"✅ Backup created: {backup_path.name}")
    
    def load_workbook(self):
        """Load the Excel workbook"""
        return load_workbook(self.config.excel_file)
    
    def get_sheet_info(self, ws: Worksheet) -> Optional[SheetInfo]:
        """Extract information about a worksheet"""
        
        # Find header row
        header_row = None
        headers = []
        
        for row in ws.iter_rows(min_row=1, max_row=20):
            values = [str(cell.value or "").strip().upper() for cell in row]
            required_headers = ["SITE ID", "CURRENT VISIT DATE", "CURRENT DG RUN HOURS", "NAME OF TECHNICIAN"]
            
            if all(any(k in v for v in values) for k in required_headers):
                header_row = row[0].row
                headers = [str(cell.value or "").strip() for cell in row]
                break
        
        if not header_row:
            return None
        
        # Normalize column names to uppercase for consistent matching
        col_map = {}
        for idx, name in enumerate(headers):
            normalized_name = name.strip().upper()
            col_map[normalized_name] = idx + 1
        
        # Find last date
        date_col = col_map.get("CURRENT VISIT DATE")
        if not date_col:
            return None
        last_date = datetime.min
        
        for r in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=r, column=date_col)
            if cell.value is None:
                continue
            
            if isinstance(cell.value, datetime):
                dt = cell.value
            else:
                try:
                    dt = datetime.strptime(str(cell.value).strip(), "%d/%m/%Y")
                except ValueError:
                    continue
            
            if dt > last_date:
                last_date = dt
        
        # Build existing keys
        site_col = col_map["SITE ID"]
        dg_col = col_map["CURRENT DG RUN HOURS"]
        date_col = col_map.get("CURRENT VISIT DATE")
        existing_keys = {
            (str(ws.cell(row=r, column=date_col).value), str(ws.cell(row=r, column=dg_col).value))
            for r in range(header_row + 1, ws.max_row + 1)
            if ws.cell(row=r, column=site_col).value
        }
        
        return SheetInfo(
            worksheet=ws,
            header_row=header_row,
            headers=headers,
            col_map=col_map,
            last_date=last_date,
            existing_keys=existing_keys
        )
    
    def find_last_data_row(self, ws: Worksheet, col_index: int) -> int:
        """Find the last row with data in a column"""
        for row in range(ws.max_row, 0, -1):
            val = ws.cell(row=row, column=col_index).value
            if val is not None and not (isinstance(val, str) and val.startswith('=')):
                return row
        return ws.max_row
    
    def ensure_column_exists(self, sheet_info: SheetInfo, column_name: str) -> SheetInfo:
        """Ensure a column exists in the sheet"""
        if column_name not in sheet_info.headers:
            logging.info(f"➕ Adding '{column_name}' column...")
            new_col_index = len(sheet_info.headers) + 1
            sheet_info.worksheet.cell(
                row=sheet_info.header_row, 
                column=new_col_index, 
                value=column_name
            )
            sheet_info.headers.append(column_name)
            sheet_info.col_map[column_name] = new_col_index
        
        return sheet_info


# ============================================================================
# DATA PROPAGATION MODULE (NEW)
# ============================================================================

@dataclass
class PropagationResult:
    """Result of data propagation operation"""
    sites_processed: int
    rows_updated: int
    dates_filled: int
    diesel_levels_filled: int
    dg_hours_filled: int = 0


class DataPropagator:
    """Propagates previous visit dates and diesel levels across site entries"""
    
    REQUIRED_COLUMNS = [
        "SITE ID",
        "PREVIOUS VISIT DATE",
        "CURRENT VISIT DATE",
        "PREVIOUS DIESEL LEVEL",
        "FUEL LEFT ON SITE",
        "PREVIOUS DG RUN HOURS",
        "CURRENT DG RUN HOURS"
    ]
    
    def __init__(self, config: Config):
        self.config = config
        self.excel_manager = ExcelManager(config)
    
    def propagate_data(self,
                      sheet_names: Optional[List[str]] = None,
                      progress_callback: Optional[Callable] = None) -> PropagationResult:
        """
        Propagate previous visit dates and diesel levels for all sites in specified sheets.
        
        Logic:
        1. For each site ID, get all rows sorted by date
        2. For first row: 
           - If PREVIOUS VISIT DATE is empty, copy CURRENT VISIT DATE
           - If PREVIOUS DIESEL LEVEL is empty, copy FUEL LEFT ON SITE
        3. For subsequent rows:
           - If PREVIOUS VISIT DATE is empty, copy CURRENT VISIT DATE from previous row
           - If PREVIOUS DIESEL LEVEL is empty, copy FUEL LEFT ON SITE from previous row
        """
        
        logging.info("🔄 Starting data propagation process...")
        self.excel_manager.backup_file()
        
        wb = self.excel_manager.load_workbook()
        
        # Determine which sheets to process
        if sheet_names is None:
            sheet_names = [self.config.fuel_sheet, self.config.ms_sheet]
        
        total_sites_processed = 0
        total_rows_updated = 0
        total_dates_filled = 0
        total_diesel_filled = 0
        total_dg_hours_filled = 0
        
        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                logging.warning(f"⚠️ Sheet '{sheet_name}' not found, skipping...")
                continue
            
            ws = wb[sheet_name]
            sheet_info = self.excel_manager.get_sheet_info(ws)
            
            if not sheet_info:
                logging.warning(f"⚠️ Could not get info for sheet '{sheet_name}', skipping...")
                continue
            
            # Ensure required columns exist - create them if missing
            sheet_info = self._ensure_required_columns(sheet_info, ws)
            
            if not self._validate_columns(sheet_info):
                logging.warning(f"⚠️ Sheet '{sheet_name}' missing required columns, skipping...")
                continue
            
            logging.info(f"📋 Processing sheet: {sheet_name}")
            
            # Process this sheet
            result = self._process_sheet(sheet_info, progress_callback)
            
            total_sites_processed += result.sites_processed
            total_rows_updated += result.rows_updated
            total_dates_filled += result.dates_filled
            total_diesel_filled += result.diesel_levels_filled
            total_dg_hours_filled += result.dg_hours_filled
            
            logging.info(f"✅ Sheet '{sheet_name}': {result.sites_processed} sites, {result.rows_updated} rows updated")
        
        # Save workbook
        wb.save(self.config.excel_file)
        logging.info(f"💾 Excel file saved successfully")
        logging.info(f"📊 Propagation Summary: {total_sites_processed} sites | {total_rows_updated} rows | "
                    f"{total_dates_filled} dates | {total_diesel_filled} diesel levels | {total_dg_hours_filled} DG hours")
        
        return PropagationResult(
            sites_processed=total_sites_processed,
            rows_updated=total_rows_updated,
            dates_filled=total_dates_filled,
            diesel_levels_filled=total_diesel_filled,
            dg_hours_filled=total_dg_hours_filled
        )
    
    def _ensure_required_columns(self, sheet_info: SheetInfo, ws: Worksheet) -> SheetInfo:
        """Ensure all required columns exist, create them if missing"""
        columns_to_add = []
        
        for col in self.REQUIRED_COLUMNS:
            if col not in sheet_info.col_map:
                columns_to_add.append(col)
        
        if columns_to_add:
            logging.info(f"➕ Adding missing columns: {columns_to_add}")
            excel_manager = ExcelManager(self.config)
            
            for col_name in columns_to_add:
                sheet_info = excel_manager.ensure_column_exists(sheet_info, col_name)
        
        return sheet_info
    
    def _validate_columns(self, sheet_info: SheetInfo) -> bool:
        """Validate that all required columns exist"""
        logging.info(f"📋 Available columns: {list(sheet_info.col_map.keys())}")
        
        missing_columns = []
        for col in self.REQUIRED_COLUMNS:
            if col not in sheet_info.col_map:
                missing_columns.append(col)
        
        if missing_columns:
            logging.error(f"❌ Missing required columns: {missing_columns}")
            logging.error(f"❌ Available columns in sheet: {list(sheet_info.col_map.keys())}")
            return False
        
        logging.info(f"✅ All required columns found")
        return True
    
    def _process_sheet(self, 
                      sheet_info: SheetInfo,
                      progress_callback: Optional[Callable] = None) -> PropagationResult:
        """Process a single sheet"""
        
        # Group rows by site ID
        site_rows = self._group_by_site(sheet_info)
        
        sites_processed = 0
        rows_updated = 0
        dates_filled = 0
        diesel_filled = 0
        dg_hours_filled = 0
        
        total_sites = len(site_rows)
        
        for site_idx, (site_id, rows) in enumerate(site_rows.items()):
            if progress_callback:
                progress = int((site_idx / total_sites) * 30) + 70  # 70-100% range
                progress_callback(progress)
            
            # Sort rows by row number to maintain chronological order
            rows.sort()
            
            # Process this site's rows
            result = self._process_site_rows(sheet_info, site_id, rows)
            
            sites_processed += 1
            rows_updated += result['rows_updated']
            dates_filled += result['dates_filled']
            diesel_filled += result['diesel_filled']
            dg_hours_filled += result['dg_hours_filled']
            
            logging.info(f"✅ Site {site_id}: {result['rows_updated']} rows updated")
        
        return PropagationResult(
            sites_processed=sites_processed,
            rows_updated=rows_updated,
            dates_filled=dates_filled,
            diesel_levels_filled=diesel_filled,
            dg_hours_filled=dg_hours_filled
        )
    
    def _group_by_site(self, sheet_info: SheetInfo) -> Dict[str, List[int]]:
        """Group row numbers by site ID"""
        site_rows = defaultdict(list)
        
        site_col = sheet_info.col_map["SITE ID"]
        
        for row in range(sheet_info.header_row + 1, sheet_info.worksheet.max_row + 1):
            site_id = sheet_info.worksheet.cell(row=row, column=site_col).value
            
            if site_id:
                site_rows[str(site_id).strip()].append(row)
        
        return dict(site_rows)
    
    def _process_site_rows(self, 
                          sheet_info: SheetInfo, 
                          site_id: str, 
                          rows: List[int]) -> Dict[str, int]:
        """Process all rows for a specific site"""
        
        ws = sheet_info.worksheet
        col_map = sheet_info.col_map
        
        try:
            prev_date_col = col_map["PREVIOUS VISIT DATE"]
            curr_date_col = col_map["CURRENT VISIT DATE"]
            prev_diesel_col = col_map["PREVIOUS DIESEL LEVEL"]
            fuel_left_col = col_map["FUEL LEFT ON SITE"]
            prev_dg_col = col_map["PREVIOUS DG RUN HOURS"]
            curr_dg_col = col_map["CURRENT DG RUN HOURS"]
        except KeyError as e:
            logging.error(f"❌ Column not found for site {site_id}: {e}")
            logging.error(f"Available columns: {list(col_map.keys())}")
            return {'rows_updated': 0, 'dates_filled': 0, 'diesel_filled': 0, 'dg_hours_filled': 0}
        
        rows_updated = 0
        dates_filled = 0
        diesel_filled = 0
        dg_hours_filled = 0
        
        previous_current_date = None
        previous_fuel_left = None
        previous_current_dg = None
        
        for row in rows:
            row_modified = False
            
            # Get current row values
            prev_date = ws.cell(row=row, column=prev_date_col).value
            curr_date = ws.cell(row=row, column=curr_date_col).value
            prev_diesel = ws.cell(row=row, column=prev_diesel_col).value
            fuel_left = ws.cell(row=row, column=fuel_left_col).value
            prev_dg = ws.cell(row=row, column=prev_dg_col).value
            curr_dg = ws.cell(row=row, column=curr_dg_col).value
            
            # Handle PREVIOUS VISIT DATE
            if self._is_empty(prev_date):
                if previous_current_date is None:
                    # First row: copy from CURRENT VISIT DATE
                    if not self._is_empty(curr_date):
                        ws.cell(row=row, column=prev_date_col, value=curr_date)
                        dates_filled += 1
                        row_modified = True
                        logging.debug(f"  Row {row}: Filled PREVIOUS VISIT DATE from CURRENT VISIT DATE")
                else:
                    # Subsequent rows: copy from previous row's CURRENT VISIT DATE
                    ws.cell(row=row, column=prev_date_col, value=previous_current_date)
                    dates_filled += 1
                    row_modified = True
                    logging.debug(f"  Row {row}: Filled PREVIOUS VISIT DATE from previous row")
            
            # Handle PREVIOUS DIESEL LEVEL
            if self._is_empty(prev_diesel):
                if previous_fuel_left is None:
                    # First row: copy from FUEL LEFT ON SITE
                    if not self._is_empty(fuel_left):
                        ws.cell(row=row, column=prev_diesel_col, value=fuel_left)
                        diesel_filled += 1
                        row_modified = True
                        logging.debug(f"  Row {row}: Filled PREVIOUS DIESEL LEVEL from FUEL LEFT ON SITE")
                else:
                    # Subsequent rows: copy from previous row's FUEL LEFT ON SITE
                    ws.cell(row=row, column=prev_diesel_col, value=previous_fuel_left)
                    diesel_filled += 1
                    row_modified = True
                    logging.debug(f"  Row {row}: Filled PREVIOUS DIESEL LEVEL from previous row")
            
            # Handle PREVIOUS DG RUN HOURS
            if self._is_empty(prev_dg):
                if previous_current_dg is None:
                    # First row: copy from CURRENT DG RUN HOURS
                    if not self._is_empty(curr_dg):
                        ws.cell(row=row, column=prev_dg_col, value=curr_dg)
                        dg_hours_filled += 1
                        row_modified = True
                        logging.debug(f"  Row {row}: Filled PREVIOUS DG RUN HOURS from CURRENT DG RUN HOURS")
                else:
                    # Subsequent rows: copy from previous row's CURRENT DG RUN HOURS
                    ws.cell(row=row, column=prev_dg_col, value=previous_current_dg)
                    dg_hours_filled += 1
                    row_modified = True
                    logging.debug(f"  Row {row}: Filled PREVIOUS DG RUN HOURS from previous row")
            
            # Update tracking variables for next iteration
            if not self._is_empty(curr_date):
                previous_current_date = curr_date
            
            if not self._is_empty(fuel_left):
                previous_fuel_left = fuel_left
            
            if not self._is_empty(curr_dg):
                previous_current_dg = curr_dg
            
            if row_modified:
                rows_updated += 1
        
        return {
            'rows_updated': rows_updated,
            'dates_filled': dates_filled,
            'diesel_filled': diesel_filled,
            'dg_hours_filled': dg_hours_filled
        }
    
    def _is_empty(self, value) -> bool:
        """Check if a cell value is empty"""
        if value is None:
            return True
        if isinstance(value, str) and value.strip() == "":
            return True
        return False


# ============================================================================
# DATA UPDATER MODULE
# ============================================================================

@dataclass
class UpdateResult:
    """Result of Excel update operation"""
    added: int
    faulty: int
    skipped: int


class DataUpdater:
    """Updates Excel sheets with parsed data"""
    
    def __init__(self, config: Config):
        self.config = config
        self.excel_manager = ExcelManager(config)
    
    def update_excel(self,
                     entries: List[Dict],
                     error_logger: Optional[Callable] = None,
                     progress_callback: Optional[Callable] = None) -> UpdateResult:
        """Update Excel file with new entries"""
        
        if not entries:
            logging.warning("⚠️ No entries to add.")
            return UpdateResult(added=0, faulty=0, skipped=0)
        
        self.excel_manager.backup_file()
        wb = self.excel_manager.load_workbook()
        
        # Load sheet information
        sheet_cache = {}
        for sheet_name in [self.config.fuel_sheet, self.config.ms_sheet]:
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            info = self.excel_manager.get_sheet_info(ws)
            
            if info:
                info = self.excel_manager.ensure_column_exists(info, "NAME OF TECHNICIAN")
                sheet_cache[sheet_name] = info
        
        if not sheet_cache:
            logging.error("❌ No valid sheets found in Excel file.")
            return UpdateResult(added=0, faulty=0, skipped=0)
        
        # Process entries
        added = 0
        faulty = 0
        skipped = 0
        total_entries = len(entries)
        
        for entry_idx, entry in enumerate(entries):
            if progress_callback:
                progress = int((entry_idx / total_entries) * 30) + 60
                progress_callback(progress)
            
            result = self._process_entry(entry, sheet_cache, error_logger)
            
            if result == 'added':
                added += 1
            elif result == 'faulty':
                faulty += 1
            elif result == 'skipped':
                skipped += 1
        
        wb.save(self.config.excel_file)
        logging.info(f"💾 Excel file saved successfully")
        logging.info(f"📊 Final Summary: {added} added | {faulty} faulty | {skipped} skipped")
        
        return UpdateResult(added=added, faulty=faulty, skipped=skipped)
    
    def _process_entry(self, 
                       entry: Dict, 
                       sheet_cache: Dict[str, SheetInfo],
                       error_logger: Optional[Callable]) -> str:
        """Process a single entry"""
        
        site_id = entry.get("SITE ID", "Unknown")
        
        # Determine target sheet
        sheet_name = self.config.ms_sheet if site_id.startswith("T3") else self.config.fuel_sheet
        
        if sheet_name not in sheet_cache:
            if error_logger:
                error_logger(f"❌ Skipped {site_id}: Target sheet '{sheet_name}' not found in workbook")
            return 'skipped'
        
        info = sheet_cache[sheet_name]
        
        # Validate date - but with more flexibility
        if "CURRENT VISIT DATE" not in entry or not entry.get("CURRENT VISIT DATE"):
            if error_logger:
                error_logger(f"⚠️ Entry for {site_id}: No date provided, but allowing entry")
            # Don't skip - we're allowing entries without dates now
        else:
            try:
                entry_date = datetime.strptime(entry["CURRENT VISIT DATE"], "%d/%m/%Y")
                if entry_date <= info.last_date:
                    if error_logger:
                        error_logger(f"❌ Skipped {site_id}: Date {entry['CURRENT VISIT DATE']} is before or equal to last date {info.last_date.strftime('%d/%m/%Y')}")
                    return 'skipped'
            except ValueError:
                if error_logger:
                    error_logger(f"❌ Skipped {site_id}: Invalid date format '{entry.get('CURRENT VISIT DATE')}'")
                return 'skipped'
        
        # Check for faulty entries
        dg_val = str(entry.get("CURRENT DG RUN HOURS", ""))
        is_faulty = dg_val == "FAULTY"
        
        # Check for duplicates - now based on Date AND Runtime
        key = (entry.get("CURRENT VISIT DATE", ""), dg_val)
        if key in info.existing_keys:
            if error_logger:
                error_logger(f"❌ Skipped {site_id}: Duplicate entry (same Date: {entry.get('CURRENT VISIT DATE', 'N/A')} and Runtime: {dg_val})")
            return 'skipped'
        
        # Write row
        row_data = [entry.get(col, "") for col in info.headers]
        last_data_row = self.excel_manager.find_last_data_row(info.worksheet, info.col_map["SITE ID"])
        start_row = last_data_row + 1
        
        for j, value in enumerate(row_data):
            info.worksheet.cell(row=start_row, column=j + 1, value=value)
        
        info.existing_keys.add(key)
        logging.info(f"✅ Added: {site_id} to {sheet_name}")
        
        return 'faulty' if is_faulty else 'added'


# ============================================================================
# UTILITIES MODULE
# ============================================================================

class UnmatchedBlockExporter:
    """Exports unmatched blocks to CSV"""
    
    def __init__(self, config: Config):
        self.config = config
        self.output_dir = config.unmatched_dir
        self.output_dir.mkdir(exist_ok=True)
    
    def export(self, unmatched_blocks: List[Dict]):
        """Export unmatched blocks to CSV"""
        if not unmatched_blocks:
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_path = self.output_dir / f"unmatched_blocks_{timestamp}.csv"
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['Block Number', 'Reason', 'Preview (first 100 chars)', 'Full Content'])
            
            for idx, block_info in enumerate(unmatched_blocks, 1):
                reason = block_info['reason']
                content = block_info['content']
                preview = content[:100].replace('\n', ' ')
                writer.writerow([idx, reason, preview, content])
        
        logging.info(f"📄 Unmatched blocks saved to: {csv_path.name}")


class GUILogger(logging.Handler):
    """Custom logging handler for GUI text widget"""
    
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget
    
    def emit(self, record):
        msg = self.format(record)
        self.text_widget.configure(state='normal')
        self.text_widget.insert(tk.END, msg + "\n")
        self.text_widget.see(tk.END)
        self.text_widget.configure(state='disabled')


# ============================================================================
# APPLICATION CONTROLLER
# ============================================================================

class AutomationController:
    """Main controller for the automation process"""
    
    def __init__(self, config: Config):
        self.config = config
        self.parser = MessageParser(config)
        self.updater = DataUpdater(config)
        self.propagator = DataPropagator(config)
        self.exporter = UnmatchedBlockExporter(config)
    
    def run(self, 
            chat_file: str,
            error_logger: Optional[Callable] = None,
            progress_callback: Optional[Callable] = None) -> Tuple[UpdateResult, int]:
        """Run the complete automation process"""
        
        # Parse messages
        logging.info("⏳ Parsing WhatsApp messages...")
        parse_result = self.parser.parse_file(chat_file, error_logger, progress_callback)
        
        if not parse_result.entries:
            logging.warning("⚠️ No valid entries found to process")
            return UpdateResult(added=0, faulty=0, skipped=0), parse_result.skipped_count
        
        # Export unmatched blocks
        if parse_result.unmatched_blocks:
            self.exporter.export(parse_result.unmatched_blocks)
        
        # Update Excel
        logging.info("💾 Updating Excel file...")
        update_result = self.updater.update_excel(
            parse_result.entries, 
            error_logger, 
            progress_callback
        )
        
        return update_result, parse_result.skipped_count
    
    def run_propagation(self,
                       progress_callback: Optional[Callable] = None) -> PropagationResult:
        """Run data propagation process"""
        logging.info("🔄 Starting data propagation...")
        return self.propagator.propagate_data(progress_callback=progress_callback)


# ============================================================================
# GUI MODULE
# ============================================================================

class AutomationGUI:
    """GUI for the automation tool"""
    
    def __init__(self, root: tk.Tk, config: Config):
        self.root = root
        self.config = config
        self.controller = AutomationController(config)
        self.chat_file = None
        self.start_time = None
        
        self._setup_ui()
        self._setup_logging()
    
    def _setup_ui(self):
        """Setup the user interface"""
        self.root.title("WhatsApp to Excel Automation - Modular v2.1")
        self.root.geometry("1100x850")
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
        # Main log area
        log_frame = ttk.LabelFrame(self.root, text="Processing Log", padding=10)
        log_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        log_frame.grid_rowconfigure(0, weight=1)
        log_frame.grid_columnconfigure(0, weight=1)
        
        self.text_area = scrolledtext.ScrolledText(
            log_frame, state='disabled', font=("Consolas", 9),
            bg="#1e1e1e", fg="#00ff00", insertbackground="white"
        )
        self.text_area.grid(row=0, column=0, sticky="nsew")
        
        # Error area
        error_frame = ttk.LabelFrame(self.root, text="Errors & Warnings", padding=10)
        error_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        error_frame.grid_rowconfigure(0, weight=1)
        error_frame.grid_columnconfigure(0, weight=1)
        
        self.error_area = scrolledtext.ScrolledText(
            error_frame, state='disabled', font=("Consolas", 9),
            height=8, bg="#2d1f1f", fg="#ff6b6b"
        )
        self.error_area.grid(row=0, column=0, sticky="nsew")
        
        # Status and progress
        status_frame = ttk.Frame(self.root)
        status_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=5)
        
        self.status_var = tk.StringVar(value="Select Excel and Chat files")
        self.eta_var = tk.StringVar(value="")
        
        ttk.Label(status_frame, textvariable=self.status_var, font=("Arial", 11, "bold")).pack(side='left', padx=10)
        ttk.Label(status_frame, textvariable=self.eta_var, font=("Arial", 10), foreground="blue").pack(side='right', padx=10)
        
        self.progress = ttk.Progressbar(self.root, length=1050, mode="determinate")
        self.progress.grid(row=3, column=0, pady=5, padx=10)
        
        # Buttons
        btn_frame = ttk.Frame(self.root)
        btn_frame.grid(row=4, column=0, pady=15)
        
        self.select_excel_btn = ttk.Button(btn_frame, text="📊 Select Excel File", command=self.select_excel_file)
        self.select_excel_btn.pack(side='left', padx=10)
        
        self.select_btn = ttk.Button(btn_frame, text="📁 Select Chat File", command=self.select_file)
        self.select_btn.pack(side='left', padx=10)
        
        self.start_btn = ttk.Button(btn_frame, text="▶️ Start Automation", command=self.start_automation_thread, state='disabled')
        self.start_btn.pack(side='left', padx=10)
        
        # NEW: Data Propagation Button
        self.propagate_btn = ttk.Button(btn_frame, text="🔄 Propagate Data", command=self.start_propagation_thread)
        self.propagate_btn.pack(side='left', padx=10)
        
        self.clear_btn = ttk.Button(btn_frame, text="🗑️ Clear Logs", command=self.clear_logs)
        self.clear_btn.pack(side='left', padx=10)
        
        # Info label
        info_text = "NEW: Data Propagation • Auto-fill previous dates, diesel levels & DG hours"
        ttk.Label(self.root, text=info_text, font=("Arial", 8), foreground="blue").grid(row=5, column=0, pady=5)
    
    def _setup_logging(self):
        """Setup logging configuration"""
        logger = logging.getLogger()
        logger.setLevel(logging.INFO)
        handler = GUILogger(self.text_area)
        handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%H:%M:%S'))
        logger.addHandler(handler)
    
    def clear_logs(self):
        """Clear all log areas"""
        self.text_area.configure(state='normal')
        self.text_area.delete('1.0', tk.END)
        self.text_area.configure(state='disabled')
        self.error_area.configure(state='normal')
        self.error_area.delete('1.0', tk.END)
        self.error_area.configure(state='disabled')
        logging.info("🧹 Logs cleared")
    
    def log_error(self, msg: str):
        """Log error message to error area"""
        self.error_area.configure(state='normal')
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.error_area.insert(tk.END, f"[{timestamp}] {msg}\n")
        self.error_area.see(tk.END)
        self.error_area.configure(state='disabled')
    
    def select_excel_file(self):
        """Handle Excel file selection"""
        path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if path:
            self.config.excel_file = path
            filename = os.path.basename(path)
            logging.info(f"📊 Excel file selected: {filename}")
            self.config.backup_dir.mkdir(exist_ok=True)
            self.config.unmatched_dir.mkdir(exist_ok=True)
            if self.chat_file:
                self.start_btn.config(state='normal')
                self.status_var.set("Ready to start")
            else:
                self.status_var.set("Now select chat file")
    
    def select_file(self):
        """Handle file selection"""
        path = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")])
        if path:
            self.chat_file = path
            filename = os.path.basename(path)
            logging.info(f"📂 Chat file selected: {filename}")
            if self.config.excel_file:
                self.start_btn.config(state='normal')
                self.status_var.set("Ready to start")
            else:
                self.status_var.set("Now select Excel file")
    
    def update_progress(self, value: int):
        """Update progress bar and ETA"""
        self.progress['value'] = value
        self.root.update_idletasks()
        
        if self.start_time and value > 0:
            elapsed = (datetime.now() - self.start_time).total_seconds()
            if value < 100:
                total_estimated = (elapsed / value) * 100
                remaining = total_estimated - elapsed
                minutes = int(remaining // 60)
                seconds = int(remaining % 60)
                self.eta_var.set(f"ETA: {minutes}m {seconds}s")
            else:
                self.eta_var.set("Complete!")
    
    def run_automation(self):
        """Run the automation process"""
        if not self.chat_file:
            logging.warning("⚠️ No chat file selected!")
            self.status_var.set("Select chat file")
            return
        
        if not self.config.excel_file:
            logging.warning("⚠️ No Excel file selected!")
            self.status_var.set("Select Excel file")
            return
        
        try:
            self.start_time = datetime.now()
            self.status_var.set("⏳ Processing...")
            self.update_progress(10)
            
            update_result, skipped_parse = self.controller.run(
                self.chat_file,
                error_logger=self.log_error,
                progress_callback=self.update_progress
            )
            
            self.status_var.set("✅ Automation Complete!")
            self.update_progress(100)
            
            # Show summary
            elapsed = (datetime.now() - self.start_time).total_seconds()
            summary = (
                f"✅ Automation Complete!\n\n"
                f"📊 Results:\n"
                f"  • New entries added: {update_result.added}\n"
                f"  • Faulty entries: {update_result.faulty}\n"
                f"  • Skipped entries: {skipped_parse + update_result.skipped}\n\n"
                f"📁 Files:\n"
                f"  • Backup: {self.config.backup_dir}/\n"
            )
            
            if skipped_parse > 0:
                summary += f"  • Unmatched blocks CSV: {self.config.unmatched_dir}/\n"
            
            summary += f"\n⏱️ Time taken: {int(elapsed)}s"
            
            messagebox.showinfo("Success", summary)
            
        except FileNotFoundError:
            logging.error(f"❌ Excel file not found: {self.config.excel_file}")
            self.status_var.set("Error: Excel file not found")
            messagebox.showerror("File Error", f"Excel file not found:\n{self.config.excel_file}")
        except Exception as e:
            logging.error(f"❌ Unexpected error: {str(e)}")
            self.status_var.set("Error occurred!")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
        finally:
            self.update_progress(0)
            self.eta_var.set("")
            self.select_btn.config(state='normal')
            self.start_btn.config(state='normal')
            self.propagate_btn.config(state='normal')
    
    def run_propagation(self):
        """Run the data propagation process"""
        try:
            self.start_time = datetime.now()
            self.status_var.set("🔄 Propagating data...")
            self.update_progress(10)
            
            result = self.controller.run_propagation(progress_callback=self.update_progress)
            
            self.status_var.set("✅ Propagation Complete!")
            self.update_progress(100)
            
            # Show summary
            elapsed = (datetime.now() - self.start_time).total_seconds()
            summary = (
                f"✅ Data Propagation Complete!\n\n"
                f"📊 Results:\n"
                f"  • Sites processed: {result.sites_processed}\n"
                f"  • Rows updated: {result.rows_updated}\n"
                f"  • Previous dates filled: {result.dates_filled}\n"
                f"  • Previous diesel levels filled: {result.diesel_levels_filled}\n"
                f"  • Previous DG run hours filled: {result.dg_hours_filled}\n\n"
                f"📁 Backup created in: {self.config.backup_dir}/\n"
                f"\n⏱️ Time taken: {int(elapsed)}s"
            )
            
            messagebox.showinfo("Success", summary)
            
        except FileNotFoundError:
            logging.error(f"❌ Excel file not found: {self.config.excel_file}")
            self.status_var.set("Error: Excel file not found")
            messagebox.showerror("File Error", f"Excel file not found:\n{self.config.excel_file}")
        except Exception as e:
            logging.error(f"❌ Unexpected error: {str(e)}")
            self.status_var.set("Error occurred!")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
        finally:
            self.update_progress(0)
            self.eta_var.set("")
            self.select_btn.config(state='normal')
            self.start_btn.config(state='normal')
            self.propagate_btn.config(state='normal')
    
    def start_automation_thread(self):
        """Start automation in separate thread"""
        self.clear_logs()
        self.status_var.set("⏳ Starting automation...")
        self.eta_var.set("Calculating...")
        self.select_btn.config(state='disabled')
        self.start_btn.config(state='disabled')
        self.propagate_btn.config(state='disabled')
        
        threading.Thread(target=self.run_automation, daemon=True).start()
    
    def start_propagation_thread(self):
        """Start data propagation in separate thread"""
        # Confirm action
        confirm = messagebox.askyesno(
            "Confirm Data Propagation",
            "This will automatically fill missing:\n"
            "• Previous Visit Dates\n"
            "• Previous Diesel Levels\n"
            "• Previous DG Run Hours\n\n"
            "For all sites in the Excel file.\n\n"
            "A backup will be created first.\n\n"
            "Continue?"
        )
        
        if not confirm:
            return
        
        self.clear_logs()
        self.status_var.set("⏳ Starting data propagation...")
        self.eta_var.set("Calculating...")
        self.select_btn.config(state='disabled')
        self.start_btn.config(state='disabled')
        self.propagate_btn.config(state='disabled')
        
        threading.Thread(target=self.run_propagation, daemon=True).start()


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main():
    """Main entry point"""
    # Load configuration
    config = Config()
    
    # Ensure directories exist
    config.backup_dir.mkdir(exist_ok=True)
    config.unmatched_dir.mkdir(exist_ok=True)
    
    # Create and run GUI
    root = tk.Tk()
    style = ttk.Style()
    style.theme_use('clam')
    
    app = AutomationGUI(root, config)
    root.mainloop()


if __name__ == "__main__":
    main()

import os
import re
import csv
import shutil
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Callable
from dataclasses import dataclass, field
import threading
import tkinter as tk
from tkinter import filedialog, scrolledtext, ttk, messagebox
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ============================================================================
# CONFIGURATION MODULE
# ============================================================================

@dataclass
class Config:
    """Application configuration"""
    excel_file: str = ""
    fuel_sheet: str = "fuel capture"
    ms_sheet: str = "MS CAPTURE"
    
    variation_mapping: Dict[str, List[str]] = field(default_factory=lambda: {
        "CURRENT DG RUN HOURS": ["Rt", "DG Current Run Time", "GD Run Time", "GD Run", "Runtime", "runtime"],
        "PREVIOUS DG RUN HOURS": ["Pre Rt", "Previous Run Time", "Previous", "Previous Run time", "Previous runtime"],
        "FUEL FOUND": ["Fuel found", "Initial fuel level", "Initial", "Fuel Total", "initial"],
        "FUEL ADDED": ["Fuel added", "Fuel Added", "Added", "fuel added"],
        "SITE ID": ["Site ID", "cbt", "CBT", "Site id", "site id", "SiteID", "Site Id"],
        "SITE NAME": ["Site name", "Site Name", "Name", "site name"],
        "SUPPLIER": ["Fuel source", "Source"],
        "DATE": ["Date"],
        "CPH": ["CPH"],
        "NAME OF TECHNICIAN": ["Technician", "Tech"]
    })
    
    numeric_fields: List[str] = field(default_factory=lambda: ["FUEL ADDED", "FUEL FOUND", "CPH"])
    allowed_sources: List[str] = field(default_factory=lambda: ["SAHARA", "MERU", "PUMA", "CCS FUEL", "TOTAL"])
    
    @property
    def backup_dir(self) -> Path:
        if not self.excel_file:
            return Path.cwd() / "backups"
        return Path(self.excel_file).parent / "backups"
    
    @property
    def unmatched_dir(self) -> Path:
        if not self.excel_file:
            return Path.cwd() / "unmatched_blocks"
        return Path(self.excel_file).parent / "unmatched_blocks"


# ============================================================================
# TECHNICIAN MAPPING MODULE
# ============================================================================

class TechnicianMapper:
    """Handles site to technician mapping"""
    
    SITE_TECHNICIAN_MAP = {
        'IHS_CBT_222A': 'FACKSON', 'IHS_CBT_075M': 'FACKSON', 'IHS_CBT_244A': 'FACKSON',
        'IHS_CBT_108M': 'FACKSON', 'IHS_CBT_105M': 'FACKSON', 'IHS_CBT_243A': 'FACKSON',
        'IHS_CBT_074M': 'FACKSON', 'IHS_CBT_305A': 'FACKSON', 'IHS_CBT_228A': 'FACKSON',
        'IHS_CBT_356A': 'FACKSON', 'IHS_CBT_409A': 'FACKSON', 'IHS_CBT_310A': 'FACKSON',
        'IHS_CBT_317A': 'FACKSON', 'IHS_CBT_291A': 'FACKSON', 'IHS_CBT_318A': 'FACKSON',
        'IHS_CBT_146A': 'FACKSON', 'IHS_CBT_168A': 'FACKSON', 'IHS_CBT_145A': 'FACKSON',
        'IHS_CBT_244B': 'FACKSON', 'IHS_CBT_324A': 'FACKSON', 'IHS_CBT_309A': 'FACKSON',
        'IHS_CBT_103A': 'FACKSON', 'IHS_CBT_204M': 'FACKSON', 'IHS_CBT_341A': 'FACKSON',
        'IHS_CBT_152A': 'FACKSON', 'IHS_CBT_303A': 'FACKSON', 'IHS_CBT_103M': 'Jaula',
        'IHS_CBT_402A': 'FACKSON', 'IHS_CBT_205A': 'FACKSON', 'IHS_CBT_107M': 'Jaula',
        'IHS_CBT_224A': 'FACKSON', 'IHS_CBT_115M': 'Jaula', 'IHS_CBT_207M': 'FACKSON',
        'IHS_CBT_200M': 'FACKSON', 'IHS_CBT_324B': 'FACKSON', 'IHS_CBT_245A': 'KAMBOLE',
        'IHS_CBT_239A': 'FACKSON', 'IHS_CBT_300M': 'FACKSON', 'IHS_CBT_287M': 'FACKSON',
        'IHS_CBT_351A': 'Jaula', 'IHS_CBT_302A': 'FACKSON', 'IHS_CBT_227M': 'Royd',
        'IHS_CBT_118A': 'FACKSON', 'IHS_CBT_132A': 'FACKSON', 'IHS_CBT_315A': 'FACKSON',
        'IHS_CBT_191A': 'FACKSON', 'IHS_CBT_402M': 'FACKSON', 'IHS_CBT_343A': 'Royd',
        'IHS_CBT_151A': 'FACKSON', 'IHS_CBT_102M': 'FACKSON', 'IHS_CBT_212A': 'FACKSON',
        'IHS_CBT_230A': 'Royd', 'IHS_CBT_212M': 'FACKSON', 'IHS_CBT_323A': 'FACKSON',
        'IHS_CBT_274A': 'FACKSON', 'IHS_CBT_319M': 'FACKSON', 'IHS_CBT_314A': 'Royd',
        'IHS_CBT_150A': 'FACKSON', 'IHS_CBT_267A': 'SAMUEL', 'IHS_CBT_129A': 'FACKSON',
        'IHS_CBT_277M': 'FACKSON', 'IHS_CBT_233A': 'KAMBOLE', 'IHS_CBT_277A': 'Royd',
        'IHS_CBT_303M': 'FACKSON', 'IHS_CBT_259A': 'SAMUEL', 'IHS_CBT_131A': 'FACKSON',
        'IHS_CBT_189A': 'FACKSON', 'IHS_CBT_102A': 'FACKSON', 'IHS_CBT_185M': 'FACKSON',
        'IHS_CBT_295A': 'FACKSON', 'IHS_CBT_118M': 'SAMUEL', 'IHS_CBT_287A': 'FACKSON',
        'IHS_CBT_133A': 'FACKSON', 'IHS_CBT_344A': 'KAMBOLE', 'IHS_CBT_347A': 'KAMBOLE',
        'IHS_CBT_321M': 'FACKSON', 'IHS_CBT_358A': 'FACKSON', 'IHS_CBT_220A': 'Jaula',
        'IHS_CBT_355A': 'FACKSON', 'IHS_CBT_132M': 'SAMUEL', 'IHS_CBT_290A': 'FACKSON',
        'IHS_CBT_332A': 'Royd', 'IHS_CBT_324M': 'FACKSON', 'IHS_CBT_083M': 'FACKSON',
        'IHS_CBT_237A': 'FACKSON', 'IHS_CBT_297A': 'FACKSON', 'IHS_CBT_084M': 'FACKSON',
        'IHS_CBT_334A': 'SAMUEL', 'IHS_CBT_079M': 'FACKSON', 'IHS_CBT_235M': 'Jaula',
        'IHS_CBT_120M': 'Jaula', 'IHS_CBT_111M': 'Jaula', 'IHS_CBT_121M': 'Jaula',
        'IHS_CBT_242A': 'FACKSON', 'IHS_CBT_089M': 'Jaula', 'IHS_CBT_093M': 'Jaula',
        'IHS_CBT_285A': 'Jaula', 'IHS_CBT_229A': 'Jaula', 'IHS_CBT_113M': 'Jaula',
        'IHS_CBT_296A': 'Jaula', 'IHS_CBT_353A': 'FACKSON', 'IHS_CBT_269A': 'Jaula',
        'IHS_CBT_114M': 'Jaula', 'IHS_CBT_241A': 'Jaula', 'IHS_CBT_090M': 'Jaula',
        'IHS_CBT_266A': 'Jaula', 'IHS_CBT_080M': 'Jaula', 'IHS_CBT_112M': 'Jaula',
        'IHS_CBT_354A': 'Jaula', 'IHS_CBT_263A': 'Jaula', 'IHS_CBT_116M': 'Jaula',
        'IHS_CBT_217M': 'Jaula', 'IHS_CBT_015M': 'KAMBOLE', 'IHS_CBT_037M': 'KAMBOLE',
        'IHS_CBT_361A': 'KAMBOLE', 'IHS_CBT_025M': 'KAMBOLE', 'IHS_CBT_023M': 'KAMBOLE',
        'IHS_CBT_222M': 'KAMBOLE', 'IHS_CBT_252A': 'KAMBOLE', 'IHS_CBT_364A': 'KAMBOLE',
        'IHS_CBT_264A': 'KAMBOLE', 'IHS_CBT_017M': 'KAMBOLE', 'IHS_CBT_218A': 'KAMBOLE',
        'IHS_CBT_226M': 'KAMBOLE', 'IHS_CBT_031M': 'KAMBOLE', 'IHS_CBT_018M': 'KAMBOLE',
        'IHS_CBT_013M': 'KAMBOLE', 'IHS_CBT_219M': 'KAMBOLE', 'IHS_CBT_257A': 'KAMBOLE',
        'IHS_CBT_029M': 'KAMBOLE', 'IHS_CBT_005M': 'KAMBOLE', 'IHS_CBT_306A': 'Royd',
        'IHS_CBT_014M': 'Royd', 'IHS_CBT_021M': 'Royd', 'IHS_CBT_085M': 'Royd',
        'IHS_CBT_010M': 'Royd', 'IHS_CBT_035M': 'Royd', 'IHS_CBT_027M': 'Royd',
        'IHS_CBT_022M': 'Royd', 'IHS_CBT_006M': 'Royd', 'IHS_CBT_024M': 'Royd',
        'IHS_CBT_033M': 'Royd', 'IHS_CBT_336A': 'Royd', 'IHS_CBT_234A': 'Royd',
        'IHS_CBT_253A': 'Royd', 'IHS_CBT_016M': 'Royd', 'IHS_CBT_349A': 'Royd',
        'IHS_CBT_091M': 'Royd', 'IHS_CBT_232A': 'Royd', 'IHS_CBT_012M': 'Royd',
        'IHS_CBT_236A': 'Royd', 'IHS_CBT_008M': 'Royd', 'IHS_CBT_056M': 'Royd',
        'IHS_CBT_293A': 'Royd', 'IHS_CBT_020M': 'Royd', 'IHS_CBT_235A': 'Royd',
        'IHS_CBT_288A': 'Royd', 'IHS_CBT_346A': 'Royd', 'IHS_CBT_040M': 'Royd',
        'IHS_CBT_003M': 'Royd', 'IHS_CBT_312A': 'Royd', 'IHS_CBT_048M': 'Royd',
        'IHS_CBT_064M': 'Royd', 'IHS_CBT_052M': 'Royd', 'IHS_CBT_062M': 'Royd',
        'IHS_CBT_058M': 'Royd', 'IHS_CBT_049M': 'Royd', 'IHS_CBT_054M': 'Royd',
        'IHS_CBT_146M': 'SAMUEL', 'IHS_CBT_126M': 'SAMUEL', 'IHS_CBT_301A': 'SAMUEL',
        'IHS_CBT_088M': 'SAMUEL', 'IHS_CBT_094M': 'FACKSON', 'IHS_CBT_342A': 'FACKSON',
        'IHS_CBT_133M': 'SAMUEL', 'IHS_CBT_261A': 'SAMUEL', 'IHS_CBT_300A': 'FACKSON',
        'IHS_CBT_073M': 'FACKSON', 'IHS_CBT_368A': 'FACKSON', 'IHS_CBT_218M': 'FACKSON',
        'IHS_CBT_276A': 'Jaula', 'IHS_CBT_123M': 'Jaula', 'IHS_CBT_216M': 'KAMBOLE',
        'IHS_CBT_335A': 'KAMBOLE', 'IHS_CBT_233M': 'Royd', 'IHS_CBT_179M': 'Royd',
        'IHS_CBT_047M': 'Royd', 'IHS_CBT_210A': 'Royd', 'IHS_CBT_050M': 'Royd',
        'IHS_CBT_307A': 'Royd', 'IHS_CTR_271A': 'Royd', 'IHS_CBT_265A': 'Royd',
        'IHS_CBT_172M': 'Royd', 'IHS_CBT_159M': 'SAMUEL', 'IHS_CBT_117M': 'SAMUEL',
        'IHS_CBT_134M': 'SAMUEL', 'IHS_CBT_167M': 'SAMUEL', 'IHS_CBT_119M': 'SAMUEL',
        'IHS_CBT_206A': 'SAMUEL', 'IHS_CBT_127M': 'SAMUEL', 'IHS_CBT_238M': 'SAMUEL',
        'IHS_CBT_155M': 'SAMUEL', 'IHS_CBT_087M': 'Jaula', 'IHS_CBT_129M': 'SAMUEL',
        'IHS_CBT_078M': 'FACKSON', 'IHS_CBT_352A': 'FACKSON', 'IHS_CBT_098M': 'Jaula',
        'IHS_CBT_333A': 'Jaula', 'IHS_CBT_099M': 'Jaula', 'IHS_CBT_002M': 'Royd',
        'IHS_CBT_207A': 'Royd', 'IHS_CBT_009M': 'Royd', 'IHS_CBT_001M': 'Royd',
        'IHS_CBT_360A': 'Royd', 'IHS_CBT_004M': 'Royd', 'IHS_CBT_043M': 'Royd',
        'IHS_CBT_044M': 'Royd', 'IHS_CBT_181M': 'Royd', 'IHS_CBT_161M': 'Royd',
        'IHS_CBT_215M': 'Royd', 'IHS_CBT_081M': 'SAMUEL', 'IHS_CBT_086M': 'SAMUEL',
        'IHS_CBT_092M': 'FACKSON', 'IHS_CBT_241M': 'SAMUEL', 'IHS_CBT_061M': 'SAMUEL',
        'IHS_CBT_345A': 'FACKSON', 'IHS_CBT_250A': 'FACKSON', 'IHS_CBT_223M': 'FACKSON',
        'IHS_CBT_217A': 'SAMUEL', 'IHS_CBT_082M': 'FACKSON', 'IHS_CBT_359A': 'KAMBOLE',
        'IHS_CBT_041M': 'KAMBOLE', 'IHS_CBT_072M': 'KAMBOLE', 'IHS_CBT_156M': 'Royd',
        'IHS_CBT_045M': 'Royd', 'IHS_CBT_283A': 'Royd', 'IHS_CBT_363A': 'Royd',
        'IHS_CBT_338A': 'Royd', 'IHS_CBT_034M': 'Royd', 'IHS_CBT_224M': 'Royd',
        'IHS_CBT_289A': 'SAMUEL', 'IHS_CBT_065M': 'Royd', 'IHS_CBT_203M': 'SAMUEL',
        'IHS_CBT_104M': 'SAMUEL', 'IHS_CBT_223A': 'SAMUEL', 'IHS_CBT_299A': 'FACKSON',
        'IHS_CBT_325A': 'Royd', 'IHS_CBT_362A': 'Royd', 'IHS_CBT_331A': 'Royd',
        'IHS_CBT_067M': 'Royd', 'IHS_CBT_011M': 'KAMBOLE', 'IHS_CBT_220M': 'Royd',
        'IHS_CBT_070M': 'Royd', 'IHS_CBT_320A': 'Royd', 'IHS_CBT_036M': 'Royd',
        'IHS_CBT_225M': 'FACKSON', 'IHS_CBT_211A': 'Royd', 'IHS_CBT_068M': 'Royd',
        'IHS_CBT_069M': 'Royd', 'IHS_CBT_193M': 'Royd', 'IHS_CBT_339A': 'Royd',
        'IHS_CBT_168M': 'Royd', 'IHS_CBT_319A': 'Royd', 'IHS_CBT_137M': 'Royd',
        'IHS_CBT_330A': 'SAMUEL', 'IHS_CBT_329A': 'SAMUEL', 'IHS_CBT_213M': 'SAMUEL',
        'IHS_CBT_214A': 'Royd', 'IHS_CBT_311A': 'Royd'
    }
    
    @classmethod
    def get_technician(cls, site_id: str) -> str:
        """Get technician name for a site ID"""
        return cls.SITE_TECHNICIAN_MAP.get(site_id, "N/A")


# ============================================================================
# DATA NORMALIZATION MODULE
# ============================================================================

class DataNormalizer:
    """Handles data normalization and validation"""
    
    SITE_ID_PATTERN = re.compile(r"CBT[_\s-]?\d+[A-Z]?", re.IGNORECASE)
    DATE_PATTERN = re.compile(r'Date\s*[:\-]?\s*(\d{1,2}/\d{1,2}/\d{2,4})', re.IGNORECASE)
    
    def __init__(self, config: Config):
        self.config = config
    
    def normalize_site_id(self, text: str) -> str:
        """Extract and normalize site ID from text"""
        if not text:
            return ""
        
        match = self.SITE_ID_PATTERN.search(text.upper())
        if not match:
            return ""
        
        site_id = match.group(0).replace(" ", "_").replace("-", "_")
        
        if not site_id.startswith("IHS_"):
            site_id = f"IHS_{site_id}"
        
        site_id = site_id.replace("IHSCBT", "IHS_CBT")
        if "CBT" in site_id and "CBT_" not in site_id:
            site_id = site_id.replace("CBT", "CBT_")
        
        return site_id
    
    def normalize_date(self, text: str) -> Optional[datetime]:
        """Extract and normalize date from text"""
        if not text:
            return None
        
        match = self.DATE_PATTERN.search(text)
        if not match:
            return None
        
        date_str = match.group(1).strip()
        for fmt in ("%d/%m/%y", "%d/%m/%Y"):
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        
        return None
    
    def normalize_supplier(self, text: str) -> str:
        """Extract and normalize supplier from text"""
        if not text:
            return ""
        
        for source in self.config.allowed_sources:
            if re.search(rf"\b{source}\b", text, re.IGNORECASE):
                return source
        
        return ""
    
    def convert_numeric(self, value: str, allow_faulty: bool = False) -> Optional[any]:
        """Convert string to numeric value"""
        if not value:
            return None
        
        value = str(value).replace(",", "").strip()
        
        try:
            return float(value) if "." in value else int(value)
        except ValueError:
            return "FAULTY" if allow_faulty else None


# ============================================================================
# MESSAGE PARSER MODULE
# ============================================================================

@dataclass
class ParseResult:
    """Result of message parsing"""
    entries: List[Dict]
    skipped_count: int
    unmatched_blocks: List[Dict]


class MessageParser:
    """Parses WhatsApp messages into structured data"""
    
    REFUEL_SPLIT = re.compile(r"REFUELING TEMPLATE", re.IGNORECASE)
    
    def __init__(self, config: Config):
        self.config = config
        self.normalizer = DataNormalizer(config)
        self.technician_mapper = TechnicianMapper()
    
    def parse_file(self, 
                   file_path: str, 
                   error_logger: Optional[Callable] = None,
                   progress_callback: Optional[Callable] = None) -> ParseResult:
        """Parse WhatsApp chat file"""
        
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        
        raw_blocks = self.REFUEL_SPLIT.split(content)[1:]
        entries = []
        unmatched_blocks = []
        skipped_count = 0
        total_blocks = len(raw_blocks)
        
        for block_idx, block in enumerate(raw_blocks):
            if progress_callback:
                progress = int((block_idx / total_blocks) * 40) + 20
                progress_callback(progress)
            
            result = self._parse_block(block, block_idx)
            
            if result is None:
                skipped_count += 1
                unmatched_blocks.append({
                    'reason': 'Parse failed',
                    'content': block
                })
                if error_logger:
                    error_logger(f"❌ Block {block_idx + 1}: Parse failed")
                continue
            
            if 'error' in result:
                skipped_count += 1
                unmatched_blocks.append({
                    'reason': result['error'],
                    'content': block
                })
                if error_logger:
                    error_logger(f"❌ Block {block_idx + 1}: {result['error']}")
                continue
            
            entries.append(result)
            logging.info(f"✅ Block {block_idx + 1}: Parsed {result.get('SITE ID', 'Unknown')}")
        
        logging.info(f"📊 Summary: {len(entries)} valid entries parsed, {skipped_count} skipped")
        
        return ParseResult(
            entries=entries,
            skipped_count=skipped_count,
            unmatched_blocks=unmatched_blocks
        )
    
    def _parse_block(self, block: str, block_idx: int) -> Optional[Dict]:
        """Parse a single message block"""
        data = {}
        lines = [line.strip() for line in block.split("\n") if line.strip()]
        
        if not lines:
            return None
        
        block_text = "\n".join(lines)
        
        # Extract fields using variation mapping
        for line in lines:
            lower_line = line.lower()
            for col, variants in self.config.variation_mapping.items():
                if col in data and col != "NAME OF TECHNICIAN":
                    continue
                
                for var in variants:
                    if var.lower() in lower_line:
                        parts = re.split(r"[:\-]", line, 1)
                        value = parts[1].strip() if len(parts) > 1 else parts[0].strip()
                        data[col] = value
                        break
                
                if col in data:
                    break
        
        # Normalize site ID
        raw_site = data.get("SITE ID", "") or block_text
        normalized_site = self.normalizer.normalize_site_id(raw_site)
        
        # More flexible validation: entry is valid if it has Site ID OR Date OR Runtime
        # Normalize date
        dt = self.normalizer.normalize_date(block_text)
        
        # Check if we have at least one of: Site ID, Date, or Runtime
        has_site_id = bool(normalized_site)
        has_date = bool(dt)
        has_runtime = bool(data.get("CURRENT DG RUN HOURS"))
        
        if not (has_site_id or has_date or has_runtime):
            return {'error': 'Missing all required fields (need at least Site ID, Date, or Runtime)'}
        
        # Set Site ID (even if empty, we'll allow it now)
        data["SITE ID"] = normalized_site if normalized_site else ""
        
        # Map technician (only if we have a site ID)
        if normalized_site:
            data["NAME OF TECHNICIAN"] = self.technician_mapper.get_technician(normalized_site)
        
        # Set date (even if empty, we'll allow it now)
        if dt:
            data["CURRENT VISIT DATE"] = dt.strftime("%d/%m/%Y")
        else:
            data["CURRENT VISIT DATE"] = ""
        
        # Normalize supplier
        data["SUPPLIER"] = self.normalizer.normalize_supplier(data.get("SUPPLIER", block_text))
        data["SITE NAME"] = data.get("SITE NAME", "").strip()
        
        # Convert numeric fields
        for field in ["CURRENT DG RUN HOURS", "PREVIOUS DG RUN HOURS"]:
            if field in data:
                data[field] = self.normalizer.convert_numeric(data[field], allow_faulty=True)
        
        for field in self.config.numeric_fields:
            if field in data:
                data[field] = self.normalizer.convert_numeric(data[field])
        
        return data


# ============================================================================
# EXCEL MANAGER MODULE
# ============================================================================

@dataclass
class SheetInfo:
    """Information about an Excel sheet"""
    worksheet: Worksheet
    header_row: int
    headers: List[str]
    col_map: Dict[str, int]
    last_date: datetime
    existing_keys: set


class ExcelManager:
    """Manages Excel file operations"""
    
    def __init__(self, config: Config):
        self.config = config
        self.backup_dir = config.backup_dir
        self.backup_dir.mkdir(exist_ok=True)
    
    def backup_file(self):
        """Create a backup of the Excel file"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.backup_dir / f"backup_{timestamp}.xlsx"
        shutil.copy(self.config.excel_file, backup_path)
        logging.info(f"✅ Backup created: {backup_path.name}")
    
    def load_workbook(self):
        """Load the Excel workbook"""
        return load_workbook(self.config.excel_file)
    
    def get_sheet_info(self, ws: Worksheet) -> Optional[SheetInfo]:
        """Extract information about a worksheet"""
        
        # Find header row
        header_row = None
        headers = []
        
        for row in ws.iter_rows(min_row=1, max_row=20):
            values = [str(cell.value or "").strip().upper() for cell in row]
            required_headers = ["SITE ID", "CURRENT VISIT DATE", "CURRENT DG RUN HOURS", "NAME OF TECHNICIAN"]
            
            if all(any(k in v for v in values) for k in required_headers):
                header_row = row[0].row
                headers = [str(cell.value or "").strip() for cell in row]
                break
        
        if not header_row:
            return None
        
        col_map = {name: idx + 1 for idx, name in enumerate(headers)}
        
        # Find last date
        date_col = col_map["CURRENT VISIT DATE"]
        last_date = datetime.min
        
        for r in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=r, column=date_col)
            if cell.value is None:
                continue
            
            if isinstance(cell.value, datetime):
                dt = cell.value
            else:
                try:
                    dt = datetime.strptime(str(cell.value).strip(), "%d/%m/%Y")
                except ValueError:
                    continue
            
            if dt > last_date:
                last_date = dt
        
        # Build existing keys
        site_col = col_map["SITE ID"]
        dg_col = col_map["CURRENT DG RUN HOURS"]
        date_col = col_map.get("CURRENT VISIT DATE")
        existing_keys = {
            (str(ws.cell(row=r, column=date_col).value), str(ws.cell(row=r, column=dg_col).value))
            for r in range(header_row + 1, ws.max_row + 1)
            if ws.cell(row=r, column=site_col).value
        }
        
        return SheetInfo(
            worksheet=ws,
            header_row=header_row,
            headers=headers,
            col_map=col_map,
            last_date=last_date,
            existing_keys=existing_keys
        )
    
    def find_last_data_row(self, ws: Worksheet, col_index: int) -> int:
        """Find the last row with data in a column"""
        for row in range(ws.max_row, 0, -1):
            val = ws.cell(row=row, column=col_index).value
            if val is not None and not (isinstance(val, str) and val.startswith('=')):
                return row
        return ws.max_row
    
    def ensure_column_exists(self, sheet_info: SheetInfo, column_name: str) -> SheetInfo:
        """Ensure a column exists in the sheet"""
        if column_name not in sheet_info.headers:
            logging.info(f"➕ Adding '{column_name}' column...")
            new_col_index = len(sheet_info.headers) + 1
            sheet_info.worksheet.cell(
                row=sheet_info.header_row, 
                column=new_col_index, 
                value=column_name
            )
            sheet_info.headers.append(column_name)
            sheet_info.col_map[column_name] = new_col_index
        
        return sheet_info


# ============================================================================
# DATA UPDATER MODULE
# ============================================================================

@dataclass
class UpdateResult:
    """Result of Excel update operation"""
    added: int
    faulty: int
    skipped: int


class DataUpdater:
    """Updates Excel sheets with parsed data"""
    
    def __init__(self, config: Config):
        self.config = config
        self.excel_manager = ExcelManager(config)
    
    def update_excel(self,
                     entries: List[Dict],
                     error_logger: Optional[Callable] = None,
                     progress_callback: Optional[Callable] = None) -> UpdateResult:
        """Update Excel file with new entries"""
        
        if not entries:
            logging.warning("⚠️ No entries to add.")
            return UpdateResult(added=0, faulty=0, skipped=0)
        
        self.excel_manager.backup_file()
        wb = self.excel_manager.load_workbook()
        
        # Load sheet information
        sheet_cache = {}
        for sheet_name in [self.config.fuel_sheet, self.config.ms_sheet]:
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            info = self.excel_manager.get_sheet_info(ws)
            
            if info:
                info = self.excel_manager.ensure_column_exists(info, "NAME OF TECHNICIAN")
                sheet_cache[sheet_name] = info
        
        if not sheet_cache:
            logging.error("❌ No valid sheets found in Excel file.")
            return UpdateResult(added=0, faulty=0, skipped=0)
        
        # Process entries
        added = 0
        faulty = 0
        skipped = 0
        total_entries = len(entries)
        
        for entry_idx, entry in enumerate(entries):
            if progress_callback:
                progress = int((entry_idx / total_entries) * 30) + 60
                progress_callback(progress)
            
            result = self._process_entry(entry, sheet_cache, error_logger)
            
            if result == 'added':
                added += 1
            elif result == 'faulty':
                faulty += 1
            elif result == 'skipped':
                skipped += 1
        
        wb.save(self.config.excel_file)
        logging.info(f"💾 Excel file saved successfully")
        logging.info(f"📊 Final Summary: {added} added | {faulty} faulty | {skipped} skipped")
        
        return UpdateResult(added=added, faulty=faulty, skipped=skipped)
    
    def _process_entry(self, 
                       entry: Dict, 
                       sheet_cache: Dict[str, SheetInfo],
                       error_logger: Optional[Callable]) -> str:
        """Process a single entry"""
        
        site_id = entry.get("SITE ID", "Unknown")
        
        # Determine target sheet
        sheet_name = self.config.ms_sheet if site_id.startswith("T3") else self.config.fuel_sheet
        
        if sheet_name not in sheet_cache:
            if error_logger:
                error_logger(f"❌ Skipped {site_id}: Target sheet '{sheet_name}' not found in workbook")
            return 'skipped'
        
        info = sheet_cache[sheet_name]
        
        # Validate date - but with more flexibility
        if "CURRENT VISIT DATE" not in entry or not entry.get("CURRENT VISIT DATE"):
            if error_logger:
                error_logger(f"⚠️ Entry for {site_id}: No date provided, but allowing entry")
            # Don't skip - we're allowing entries without dates now
        else:
            try:
                entry_date = datetime.strptime(entry["CURRENT VISIT DATE"], "%d/%m/%Y")
                if entry_date <= info.last_date:
                    if error_logger:
                        error_logger(f"❌ Skipped {site_id}: Date {entry['CURRENT VISIT DATE']} is before or equal to last date {info.last_date.strftime('%d/%m/%Y')}")
                    return 'skipped'
            except ValueError:
                if error_logger:
                    error_logger(f"❌ Skipped {site_id}: Invalid date format '{entry.get('CURRENT VISIT DATE')}'")
                return 'skipped'
        
        # Check for faulty entries
        dg_val = str(entry.get("CURRENT DG RUN HOURS", ""))
        is_faulty = dg_val == "FAULTY"
        
        # Check for duplicates - now based on Date AND Runtime
        key = (entry.get("CURRENT VISIT DATE", ""), dg_val)
        if key in info.existing_keys:
            if error_logger:
                error_logger(f"❌ Skipped {site_id}: Duplicate entry (same Date: {entry.get('CURRENT VISIT DATE', 'N/A')} and Runtime: {dg_val})")
            return 'skipped'
        
        # Write row
        row_data = [entry.get(col, "") for col in info.headers]
        last_data_row = self.excel_manager.find_last_data_row(info.worksheet, info.col_map["SITE ID"])
        start_row = last_data_row + 1
        
        for j, value in enumerate(row_data):
            info.worksheet.cell(row=start_row, column=j + 1, value=value)
        
        info.existing_keys.add(key)
        logging.info(f"✅ Added: {site_id} to {sheet_name}")
        
        return 'faulty' if is_faulty else 'added'


# ============================================================================
# UTILITIES MODULE
# ============================================================================

class UnmatchedBlockExporter:
    """Exports unmatched blocks to CSV"""
    
    def __init__(self, config: Config):
        self.config = config
        self.output_dir = config.unmatched_dir
        self.output_dir.mkdir(exist_ok=True)
    
    def export(self, unmatched_blocks: List[Dict]):
        """Export unmatched blocks to CSV"""
        if not unmatched_blocks:
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_path = self.output_dir / f"unmatched_blocks_{timestamp}.csv"
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['Block Number', 'Reason', 'Preview (first 100 chars)', 'Full Content'])
            
            for idx, block_info in enumerate(unmatched_blocks, 1):
                reason = block_info['reason']
                content = block_info['content']
                preview = content[:100].replace('\n', ' ')
                writer.writerow([idx, reason, preview, content])
        
        logging.info(f"📄 Unmatched blocks saved to: {csv_path.name}")


class GUILogger(logging.Handler):
    """Custom logging handler for GUI text widget"""
    
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget
    
    def emit(self, record):
        msg = self.format(record)
        self.text_widget.configure(state='normal')
        self.text_widget.insert(tk.END, msg + "\n")
        self.text_widget.see(tk.END)
        self.text_widget.configure(state='disabled')


# ============================================================================
# APPLICATION CONTROLLER
# ============================================================================

class AutomationController:
    """Main controller for the automation process"""
    
    def __init__(self, config: Config):
        self.config = config
        self.parser = MessageParser(config)
        self.updater = DataUpdater(config)
        self.exporter = UnmatchedBlockExporter(config)
    
    def run(self, 
            chat_file: str,
            error_logger: Optional[Callable] = None,
            progress_callback: Optional[Callable] = None) -> Tuple[UpdateResult, int]:
        """Run the complete automation process"""
        
        # Parse messages
        logging.info("⏳ Parsing WhatsApp messages...")
        parse_result = self.parser.parse_file(chat_file, error_logger, progress_callback)
        
        if not parse_result.entries:
            logging.warning("⚠️ No valid entries found to process")
            return UpdateResult(added=0, faulty=0, skipped=0), parse_result.skipped_count
        
        # Export unmatched blocks
        if parse_result.unmatched_blocks:
            self.exporter.export(parse_result.unmatched_blocks)
        
        # Update Excel
        logging.info("💾 Updating Excel file...")
        update_result = self.updater.update_excel(
            parse_result.entries, 
            error_logger, 
            progress_callback
        )
        
        return update_result, parse_result.skipped_count


# ============================================================================
# GUI MODULE
# ============================================================================

class AutomationGUI:
    """GUI for the automation tool"""
    
    def __init__(self, root: tk.Tk, config: Config):
        self.root = root
        self.config = config
        self.controller = AutomationController(config)
        self.chat_file = None
        self.start_time = None
        
        self._setup_ui()
        self._setup_logging()
    
    def _setup_ui(self):
        """Setup the user interface"""
        self.root.title("WhatsApp to Excel Automation - Modular v2.0")
        self.root.geometry("1100x800")
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
        # Main log area
        log_frame = ttk.LabelFrame(self.root, text="Processing Log", padding=10)
        log_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        log_frame.grid_rowconfigure(0, weight=1)
        log_frame.grid_columnconfigure(0, weight=1)
        
        self.text_area = scrolledtext.ScrolledText(
            log_frame, state='disabled', font=("Consolas", 9),
            bg="#1e1e1e", fg="#00ff00", insertbackground="white"
        )
        self.text_area.grid(row=0, column=0, sticky="nsew")
        
        # Error area
        error_frame = ttk.LabelFrame(self.root, text="Errors & Warnings", padding=10)
        error_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        error_frame.grid_rowconfigure(0, weight=1)
        error_frame.grid_columnconfigure(0, weight=1)
        
        self.error_area = scrolledtext.ScrolledText(
            error_frame, state='disabled', font=("Consolas", 9),
            height=8, bg="#2d1f1f", fg="#ff6b6b"
        )
        self.error_area.grid(row=0, column=0, sticky="nsew")
        
        # Status and progress
        status_frame = ttk.Frame(self.root)
        status_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=5)
        
        self.status_var = tk.StringVar(value="Select Excel and Chat files")
        self.eta_var = tk.StringVar(value="")
        
        ttk.Label(status_frame, textvariable=self.status_var, font=("Arial", 11, "bold")).pack(side='left', padx=10)
        ttk.Label(status_frame, textvariable=self.eta_var, font=("Arial", 10), foreground="blue").pack(side='right', padx=10)
        
        self.progress = ttk.Progressbar(self.root, length=1050, mode="determinate")
        self.progress.grid(row=3, column=0, pady=5, padx=10)
        
        # Buttons
        btn_frame = ttk.Frame(self.root)
        btn_frame.grid(row=4, column=0, pady=15)
        
        self.select_excel_btn = ttk.Button(btn_frame, text="📊 Select Excel File", command=self.select_excel_file)
        self.select_excel_btn.pack(side='left', padx=10)
        
        self.select_btn = ttk.Button(btn_frame, text="📁 Select Chat File", command=self.select_file)
        self.select_btn.pack(side='left', padx=10)
        
        self.start_btn = ttk.Button(btn_frame, text="▶️ Start Automation", command=self.start_automation_thread, state='disabled')
        self.start_btn.pack(side='left', padx=10)
        
        self.clear_btn = ttk.Button(btn_frame, text="🗑️ Clear Logs", command=self.clear_logs)
        self.clear_btn.pack(side='left', padx=10)
        
        # Info label
        info_text = "Modular Architecture: Separated concerns • Clean code • Easy maintenance"
        ttk.Label(self.root, text=info_text, font=("Arial", 8), foreground="gray").grid(row=5, column=0, pady=5)
    
    def _setup_logging(self):
        """Setup logging configuration"""
        logger = logging.getLogger()
        logger.setLevel(logging.INFO)
        handler = GUILogger(self.text_area)
        handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%H:%M:%S'))
        logger.addHandler(handler)
    
    def clear_logs(self):
        """Clear all log areas"""
        self.text_area.configure(state='normal')
        self.text_area.delete('1.0', tk.END)
        self.text_area.configure(state='disabled')
        self.error_area.configure(state='normal')
        self.error_area.delete('1.0', tk.END)
        self.error_area.configure(state='disabled')
        logging.info("🧹 Logs cleared")
    
    def log_error(self, msg: str):
        """Log error message to error area"""
        self.error_area.configure(state='normal')
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.error_area.insert(tk.END, f"[{timestamp}] {msg}\n")
        self.error_area.see(tk.END)
        self.error_area.configure(state='disabled')
    
    def select_file(self):
        """Handle file selection"""
        path = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")])
        if path:
            self.chat_file = path
            filename = os.path.basename(path)
            logging.info(f"📂 Chat file selected: {filename}")
            if self.config.excel_file:
                self.start_btn.config(state='normal')
                self.status_var.set("Ready to start")
            else:
                self.status_var.set("Now select Excel file")
    
    def update_progress(self, value: int):
        """Update progress bar and ETA"""
        self.progress['value'] = value
        self.root.update_idletasks()
        
        if self.start_time and value > 0:
            elapsed = (datetime.now() - self.start_time).total_seconds()
            if value < 100:
                total_estimated = (elapsed / value) * 100
                remaining = total_estimated - elapsed
                minutes = int(remaining // 60)
                seconds = int(remaining % 60)
                self.eta_var.set(f"ETA: {minutes}m {seconds}s")
            else:
                self.eta_var.set("Complete!")
    
    def run_automation(self):
        """Run the automation process"""
        if not self.chat_file:
            logging.warning("⚠️ No chat file selected!")
            self.status_var.set("Select chat file")
            return
        
        if not self.config.excel_file:
            logging.warning("⚠️ No Excel file selected!")
            self.status_var.set("Select Excel file")
            return
        
        try:
            self.start_time = datetime.now()
            self.status_var.set("⏳ Processing...")
            self.update_progress(10)
            
            update_result, skipped_parse = self.controller.run(
                self.chat_file,
                error_logger=self.log_error,
                progress_callback=self.update_progress
            )
            
            self.status_var.set("✅ Automation Complete!")
            self.update_progress(100)
            
            # Show summary
            elapsed = (datetime.now() - self.start_time).total_seconds()
            summary = (
                f"✅ Automation Complete!\n\n"
                f"📊 Results:\n"
                f"  • New entries added: {update_result.added}\n"
                f"  • Faulty entries: {update_result.faulty}\n"
                f"  • Skipped entries: {skipped_parse + update_result.skipped}\n\n"
                f"📁 Files:\n"
                f"  • Backup: {self.config.backup_dir}/\n"
            )
            
            if skipped_parse > 0:
                summary += f"  • Unmatched blocks CSV: {self.config.unmatched_dir}/\n"
            
            summary += f"\n⏱️ Time taken: {int(elapsed)}s"
            
            messagebox.showinfo("Success", summary)
            
        except FileNotFoundError:
            logging.error(f"❌ Excel file not found: {self.config.excel_file}")
            self.status_var.set("Error: Excel file not found")
            messagebox.showerror("File Error", f"Excel file not found:\n{self.config.excel_file}")
        except Exception as e:
            logging.error(f"❌ Unexpected error: {str(e)}")
            self.status_var.set("Error occurred!")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
        finally:
            self.update_progress(0)
            self.eta_var.set("")
            self.select_btn.config(state='normal')
            self.start_btn.config(state='normal')
    
    def start_automation_thread(self):
        """Start automation in separate thread"""
        self.clear_logs()
        self.status_var.set("⏳ Starting automation...")
        self.eta_var.set("Calculating...")
        self.select_btn.config(state='disabled')
        self.start_btn.config(state='disabled')
        
        threading.Thread(target=self.run_automation, daemon=True).start()


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main():
    """Main entry point"""
    # Load configuration
    config = Config()
    
    # Ensure directories exist
    config.backup_dir.mkdir(exist_ok=True)
    config.unmatched_dir.mkdir(exist_ok=True)
    
    # Create and run GUI
    root = tk.Tk()
    style = ttk.Style()
    style.theme_use('clam')
    
    app = AutomationGUI(root, config)
    root.mainloop()


if __name__ == "__main__":
    main()